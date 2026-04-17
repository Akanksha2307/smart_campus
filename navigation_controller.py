from __future__ import annotations
import re
import math as _math
from typing import List, Tuple, Optional, Dict

# FIX: Module-level OpenCV availability flag (mirrors navigation_screen.py)
try:
    import cv2 as _cv2_test
    CV2_AVAILABLE = True
    del _cv2_test
except ImportError:
    CV2_AVAILABLE = False
# networkx replaced with a pure-Python graph — Android NDK compatible.
# Implements the subset of the nx.Graph API used by this module:
#   G.add_node / G.add_edge / __contains__ / nodes / edges[]
#   has_path / shortest_path / NetworkXNoPath


class _SimpleGraph:
    """Lightweight undirected weighted graph (Dijkstra-based shortest path)."""

    class _EdgeView:
        """Allows _GRAPH.edges[a, b].get('weight', 0) syntax."""
        def __init__(self, adj):
            self._adj = adj

        def __getitem__(self, key):
            a, b = key
            data = self._adj.get(a, {}).get(b)
            if data is None:
                raise KeyError(f"No edge ({a!r}, {b!r})")
            return data

    def __init__(self):
        self._nodes: dict = {}   # node_id -> attr dict
        self._adj:   dict = {}   # node_id -> {neighbour -> attr dict}

    # ── mutation ──────────────────────────────────────────────────────────────
    def add_node(self, nid, **attrs):
        if nid not in self._nodes:
            self._nodes[nid] = {}
            self._adj[nid]   = {}
        self._nodes[nid].update(attrs)

    def add_edge(self, a, b, **attrs):
        for n in (a, b):
            if n not in self._nodes:
                self.add_node(n)
        self._adj[a][b] = attrs
        self._adj[b][a] = attrs

    # ── query ─────────────────────────────────────────────────────────────────
    def __contains__(self, nid):
        return nid in self._nodes

    @property
    def nodes(self):
        return self._nodes

    @property
    def edges(self):
        return self._EdgeView(self._adj)

    def neighbors(self, nid):
        return iter(self._adj.get(nid, {}))

    def has_edge(self, a, b) -> bool:
        return b in self._adj.get(a, {})

    def __getitem__(self, nid):
        """Allow _GRAPH[a][b]['weight'] = dist  (edge-weight mutation)."""
        return self._adj[nid]

    # ── path-finding (Dijkstra) ───────────────────────────────────────────────
    def has_path(self, src, dst):
        if src not in self._nodes or dst not in self._nodes:
            return False
        visited = set()
        queue   = [src]
        while queue:
            nxt = []
            for n in queue:
                if n == dst:
                    return True
                if n in visited:
                    continue
                visited.add(n)
                nxt.extend(nb for nb in self._adj.get(n, {}) if nb not in visited)
            queue = nxt
        return False

    def shortest_path(self, src, dst, weight="weight"):
        import heapq
        if src not in self._nodes or dst not in self._nodes:
            raise _SimpleGraph.NetworkXNoPath(f"No path {src!r} → {dst!r}")
        dist   = {src: 0.0}
        prev   = {}
        heap   = [(0.0, src)]
        while heap:
            d, u = heapq.heappop(heap)
            if d > dist.get(u, float("inf")):
                continue
            if u == dst:
                path = []
                while u in prev:
                    path.append(u); u = prev[u]
                path.append(src)
                return path[::-1]
            for v, edata in self._adj.get(u, {}).items():
                w = edata.get(weight, 1) if weight else 1
                nd = d + w
                if nd < dist.get(v, float("inf")):
                    dist[v] = nd; prev[v] = u
                    heapq.heappush(heap, (nd, v))
        raise _SimpleGraph.NetworkXNoPath(f"No path {src!r} → {dst!r}")

    class NetworkXNoPath(Exception):
        pass


# Module-level aliases to match the nx.* call sites below
class nx:  # noqa: N801
    Graph           = _SimpleGraph
    NetworkXNoPath  = _SimpleGraph.NetworkXNoPath

    @staticmethod
    def has_path(G, src, dst):
        return G.has_path(src, dst)

    @staticmethod
    def shortest_path(G, src, dst, weight="weight"):
        return G.shortest_path(src, dst, weight=weight)

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

_BLOCK_LONG = {"S": "South Block", "N": "North Block",
               "E": "East Block",  "W": "West Block"}

_FLOOR_NAME = {0: "Ground Floor", 1: "First Floor",
               2: "Second Floor", 3: "Third Floor", 4: "Fourth Floor"}

# Corridor column ranges per block per floor  (min, max)
_COR_RANGES: Dict[str, Dict[int, Tuple[int, int]]] = {
    "S": {0: (2, 12), 1: (2, 13), 2: (2, 13), 3: (2, 13)},
    "N": {0: (2, 12), 1: (2, 13), 2: (2, 13), 3: (2, 13)},
    "E": {0: (2,  8), 1: (2,  8), 2: (2,  6)},
    "W": {0: (2, 10), 1: (2,  8), 2: (2,  8), 3: (2, 10), 4: (2,  8)},
}

# Floor letter used in corridor node IDs
_FC = {0: "G", 1: "F", 2: "S", 3: "T", 4: "4"}

# Weight constants (metres; used for Dijkstra)
_W_COR   = 8    # one corridor segment
_W_STAIR = 15   # one staircase floor
_W_LIFT  = 10   # one lift floor
_W_DOOR  = 5    # building doorway crossing


# ─────────────────────────────────────────────────────────────────────────────
#  HELPER – compact node-record builders
# ─────────────────────────────────────────────────────────────────────────────

def _r(name, rnum, blk, fl, fln, col, side):
    """Room node record."""
    return dict(name=name, room_num=rnum, block=blk,
                floor=fl, floor_num=fln, type="room", col=col, side=side)

def _fac(name, blk, fl, fln, col, side, ftype):
    """Facility node record (lift / stairs / washroom / entrance / landmark)."""
    return dict(name=name, room_num="", block=blk,
                floor=fl, floor_num=fln, type=ftype, col=col, side=side)


# ─────────────────────────────────────────────────────────────────────────────
#  ALL INDOOR NODES
# ─────────────────────────────────────────────────────────────────────────────

_NODES: Dict[str, dict] = {}

# ══════════════════════════════════════════════════════════════════════════════
#  SOUTH BLOCK  (S_)
# ══════════════════════════════════════════════════════════════════════════════
_B, _BN = "S", "SOUTH BLOCK"
_GF, _FF, _SF, _TF = "Ground Floor", "First Floor", "Second Floor", "Third Floor"

_NODES.update({
    # ── Ground Floor ─────────────────────────────────────────────────────────
    "S_G14":      _r("CSE-A 3rd yr",           "G14",    _B, _GF, 0, col=2,  side="top"),
    "S_G15":      _r("English & Mgmt",         "G15",    _B, _GF, 0, col=3,  side="top"),
    "S_G16":      _r("CSE-A 4th yr",           "G16",    _B, _GF, 0, col=4,  side="top"),
    "S_G_GRLSWC": _fac("Girls Washroom",       _B, _GF,  0, col=5,  side="top", ftype="washroom"),
    "S_G_STAIR":  _fac("Staircase",            _B, _GF,  0, col=6,  side="top", ftype="stairs"),
    "S_G_LIFT":   _fac("Lift",                 _B, _GF,  0, col=7,  side="top", ftype="lift"),
    "S_G_GENSWC": _fac("Gents Washroom",       _B, _GF,  0, col=8,  side="top", ftype="washroom"),
    "S_G02":      _r("Sick Room",              "G02",    _B, _GF, 0, col=9,  side="top"),
    "S_G03G04":   _r("ICS Lab-2",             "G03,G04", _B, _GF, 0, col=10, side="top"),
    "S_G05":      _r("Tutorial Room",          "G05",    _B, _GF, 0, col=11, side="top"),
    "S_G06":      _r("Staff Room",             "G06",    _B, _GF, 0, col=12, side="top"),
    "S_G13":      _r("CSE-A 2nd yr",           "G13",    _B, _GF, 0, col=2,  side="bottom"),
    "S_G12":      _r("FM & HM Lab",            "G12",    _B, _GF, 0, col=3,  side="bottom"),
    "S_G11":      _r("Locker Room",            "G11",    _B, _GF, 0, col=4,  side="bottom"),
    "S_ENTR":     _fac("South Block Entrance", _BN, _GF, 0, col=6,  side="bottom", ftype="entrance"),
    "S_G09":      _r("Power Electronics Lab",  "G09",    _B, _GF, 0, col=10, side="bottom"),
    "S_G08":      _r("CSE Project Lab-4",      "G08",    _B, _GF, 0, col=11, side="bottom"),
    "S_G07":      _r("CSE-A 1st yr",           "G07",    _B, _GF, 0, col=12, side="bottom"),
    # ── First Floor ──────────────────────────────────────────────────────────
    "S_105":      _r("Staff Room",             "105",    _B, _FF, 1, col=2,  side="top"),
    "S_104":      _r("CSE-B 4th yr",           "104",    _B, _FF, 1, col=3,  side="top"),
    "S_103":      _r("Tutorial Room",          "103",    _B, _FF, 1, col=4,  side="top"),
    "S_102":      _r("CSE HOD Cabin",          "102",    _B, _FF, 1, col=5,  side="top"),
    "S_F_GENSWC": _fac("Gents Washroom",       _B, _FF,  1, col=6,  side="top", ftype="washroom"),
    "S_F_STAIR":  _fac("Staircase",            _B, _FF,  1, col=7,  side="top", ftype="stairs"),
    "S_F_LIFT":   _fac("Lift",                 _B, _FF,  1, col=8,  side="top", ftype="lift"),
    "S_F_STAIR2": _fac("Staircase (2)",        _B, _FF,  1, col=9,  side="top", ftype="stairs"),
    "S_F_GRLSWC": _fac("Girls Washroom",       _B, _FF,  1, col=10, side="top", ftype="washroom"),
    "S_115":      _r("CSE Lab-3",              "115",    _B, _FF, 1, col=11, side="top"),
    "S_114":      _r("CSE Lab-2",              "114",    _B, _FF, 1, col=12, side="top"),
    "S_113":      _r("CSE-B 3rd yr",           "113",    _B, _FF, 1, col=13, side="top"),
    "S_106":      _r("CSE-B 1st yr",           "106",    _B, _FF, 1, col=2,  side="bottom"),
    "S_107":      _r("Call Lab-1",             "107",    _B, _FF, 1, col=3,  side="bottom"),
    "S_108":      _r("Call Lab-2",             "108",    _B, _FF, 1, col=4,  side="bottom"),
    "S_F_NOTICE": _fac("Notice Board",         _B, _FF,  1, col=5,  side="bottom", ftype="landmark"),
    "S_109":      _r("UPS Room",               "109",    _B, _FF, 1, col=10, side="bottom"),
    "S_110":      _r("ICS Lab-2",              "110",    _B, _FF, 1, col=11, side="bottom"),
    "S_111":      _r("CSE Lab-1",              "111",    _B, _FF, 1, col=12, side="bottom"),
    "S_112":      _r("CSE-B 2nd yr",           "112",    _B, _FF, 1, col=13, side="bottom"),
    # ── Second Floor ─────────────────────────────────────────────────────────
    "S_204":      _r("Staff Room",             "204",    _B, _SF, 2, col=2,  side="top"),
    "S_203":      _r("Staff Room",             "203",    _B, _SF, 2, col=3,  side="top"),
    "S_202":      _r("Normal Room",            "202",    _B, _SF, 2, col=4,  side="top"),
    "S_S_GENSWC": _fac("Gents Washroom",       _B, _SF,  2, col=5,  side="top", ftype="washroom"),
    "S_S_STAIR":  _fac("Staircase",            _B, _SF,  2, col=6,  side="top", ftype="stairs"),
    "S_S_LIFT":   _fac("Lift",                 _B, _SF,  2, col=7,  side="top", ftype="lift"),
    "S_S_STAIR2": _fac("Staircase (2)",        _B, _SF,  2, col=9,  side="top", ftype="stairs"),
    "S_S_GRLSWC": _fac("Girls Washroom",       _B, _SF,  2, col=10, side="top", ftype="washroom"),
    "S_217":      _r("IT Lab-1",               "217",    _B, _SF, 2, col=11, side="top"),
    "S_216":      _r("IT Lab-2",               "216",    _B, _SF, 2, col=12, side="top"),
    "S_215":      _r("IT-A 3rd yr",            "215",    _B, _SF, 2, col=13, side="top"),
    "S_205":      _r("IT-A 1st yr",            "205",    _B, _SF, 2, col=2,  side="bottom"),
    "S_206":      _r("GE Staff Room",          "206",    _B, _SF, 2, col=3,  side="bottom"),
    "S_207":      _r("Tutorial Room",          "207",    _B, _SF, 2, col=4,  side="bottom"),
    "S_208":      _r("IT-A 4th yr",            "208",    _B, _SF, 2, col=5,  side="bottom"),
    "S_209":      _r("Locker Room",            "209",    _B, _SF, 2, col=6,  side="bottom"),
    "S_S_NOTICE": _fac("Notice Board",         _B, _SF,  2, col=7,  side="bottom", ftype="landmark"),
    "S_210":      _r("Physics Lab-1",          "210",    _B, _SF, 2, col=9,  side="bottom"),
    "S_211":      _r("Physics Lab-2",          "211",    _B, _SF, 2, col=10, side="bottom"),
    "S_212":      _r("IT Lab-2",               "212",    _B, _SF, 2, col=11, side="bottom"),
    "S_213":      _r("IT Lab-3",               "213",    _B, _SF, 2, col=12, side="bottom"),
    "S_214":      _r("IT-A 2nd yr",            "214",    _B, _SF, 2, col=13, side="bottom"),
    # ── Third Floor ──────────────────────────────────────────────────────────
    "S_305":      _r("Staff Room",             "305",    _B, _TF, 3, col=2,  side="top"),
    "S_304":      _r("IT-B 4th yr",            "304",    _B, _TF, 3, col=3,  side="top"),
    "S_303":      _r("Tutorial Room",          "303",    _B, _TF, 3, col=4,  side="top"),
    "S_302":      _r("IT HOD Cabin",           "302",    _B, _TF, 3, col=5,  side="top"),
    "S_T_GENSWC": _fac("Gents Washroom",       _B, _TF,  3, col=6,  side="top", ftype="washroom"),
    "S_T_STAIR":  _fac("Staircase",            _B, _TF,  3, col=7,  side="top", ftype="stairs"),
    "S_T_LIFT":   _fac("Lift",                 _B, _TF,  3, col=8,  side="top", ftype="lift"),
    "S_T_STAIR2": _fac("Staircase (2)",        _B, _TF,  3, col=9,  side="top", ftype="stairs"),
    "S_T_GRLSWC": _fac("Girls Washroom",       _B, _TF,  3, col=10, side="top", ftype="washroom"),
    "S_314":      _r("IT Lab-4",               "314",    _B, _TF, 3, col=11, side="top"),
    "S_313":      _r("IT Workshop",            "313",    _B, _TF, 3, col=12, side="top"),
    "S_312":      _r("IT-B 3rd yr",            "312",    _B, _TF, 3, col=13, side="top"),
    "S_306":      _r("IT-B 1st yr",            "306",    _B, _TF, 3, col=2,  side="bottom"),
    "S_307":      _r("Drawing Hall",           "307",    _B, _TF, 3, col=3,  side="bottom"),
    "S_T_NOTICE": _fac("Notice Board",         _B, _TF,  3, col=5,  side="bottom", ftype="landmark"),
    "S_308":      _r("Chemistry Lab-1",        "308",    _B, _TF, 3, col=9,  side="bottom"),
    "S_309":      _r("Locker Room",            "309",    _B, _TF, 3, col=10, side="bottom"),
    "S_310":      _r("Chemistry Lab-2",        "310",    _B, _TF, 3, col=11, side="bottom"),
    "S_311":      _r("IT-B 2nd yr",            "311",    _B, _TF, 3, col=13, side="bottom"),
})

# ══════════════════════════════════════════════════════════════════════════════
#  NORTH BLOCK  (N_)
#  Same column layout as South Block — same room-number ranges, same grid.
# ══════════════════════════════════════════════════════════════════════════════
_B, _BN = "N", "NORTH BLOCK"

_NODES.update({
    # ── Ground Floor ─────────────────────────────────────────────────────────
    "N_G13":      _r("Staff Room",                          "G13",  _B, _GF, 0, col=2,  side="top"),
    "N_G14":      _r("Common Room",                         "G14",  _B, _GF, 0, col=3,  side="top"),
    "N_G15":      _r("Power Systems Lab",                   "G15",  _B, _GF, 0, col=4,  side="top"),
    "N_G16":      _r("Panel Board Room",                    "G16",  _B, _GF, 0, col=5,  side="top"),
    "N_G_GRLSWC": _fac("Girls Washroom",        _B, _GF,    0, col=6,  side="top", ftype="washroom"),  # FIX: was col=5, conflicted with N_G16
    "N_G_STAIR":  _fac("Staircase",             _B, _GF,    0, col=6,  side="top", ftype="stairs"),
    "N_G_LIFT":   _fac("Lift",                  _B, _GF,    0, col=7,  side="top", ftype="lift"),
    "N_G_GENSWC": _fac("Gents Washroom",        _B, _GF,    0, col=8,  side="top", ftype="washroom"),
    "N_G02":      _r("Room G02",                            "G02",  _B, _GF, 0, col=9,  side="top"),
    "N_G03":      _r("EEE HOD Cabin",                       "G03",  _B, _GF, 0, col=10, side="top"),
    "N_G04":      _r("Electrical Measurements & Instruments","G04", _B, _GF, 0, col=11, side="top"),
    "N_G05":      _r("Control Systems Lab",                 "G05",  _B, _GF, 0, col=12, side="top"),
    "N_G06":      _r("Room G06",                            "G06",  _B, _GF, 0, col=12, side="top"),
    "N_G11":      _r("Electrical Machine Lab",              "G11",  _B, _GF, 0, col=4,  side="bottom"),
    "N_G12":      _r("Room G12",                            "G12",  _B, _GF, 0, col=3,  side="bottom"),
    "N_ENTR":     _fac("North Block Entrance",  _BN, _GF,   0, col=6,  side="bottom", ftype="entrance"),
    "N_G10":      _r("Power Simulation",                    "G10",  _B, _GF, 0, col=10, side="bottom"),
    "N_G09":      _r("Room G09",                            "G09",  _B, _GF, 0, col=11, side="bottom"),
    "N_G08":      _r("Tutorial Room",                       "G08",  _B, _GF, 0, col=10, side="bottom"),
    "N_G07":      _r("Room G07",                            "G07",  _B, _GF, 0, col=12, side="bottom"),
    # ── First Floor ──────────────────────────────────────────────────────────
    "N_105":      _r("Staff Room",             "105",   _B, _FF, 1, col=2,  side="top"),
    "N_104":      _r("Room 104",               "104",   _B, _FF, 1, col=3,  side="top"),
    "N_103":      _r("Computer Lab-1",         "103",   _B, _FF, 1, col=4,  side="top"),
    "N_102":      _r("ECE HOD Cabin",          "102",   _B, _FF, 1, col=5,  side="top"),
    "N_F_GENSWC": _fac("Gents Washroom",       _B, _FF, 1, col=6,  side="top", ftype="washroom"),
    "N_F_STAIR":  _fac("Staircase",            _B, _FF, 1, col=7,  side="top", ftype="stairs"),
    "N_F_LIFT":   _fac("Lift",                 _B, _FF, 1, col=8,  side="top", ftype="lift"),
    "N_F_STAIR2": _fac("Staircase (2)",        _B, _FF, 1, col=9,  side="top", ftype="stairs"),
    "N_F_GRLSWC": _fac("Girls Washroom",       _B, _FF, 1, col=10, side="top", ftype="washroom"),
    "N_115":      _r("CSE Lab-3",              "115",   _B, _FF, 1, col=11, side="top"),
    "N_117":      _r("Microwave & Optical Lab","117",   _B, _FF, 1, col=12, side="top"),
    "N_112":      _r("Staff Room",             "112",   _B, _FF, 1, col=13, side="top"),
    "N_106":      _r("Embedded System Lab",    "106",   _B, _FF, 1, col=2,  side="bottom"),
    "N_107":      _r("Computer Lab-2",         "107",   _B, _FF, 1, col=3,  side="bottom"),
    "N_108":      _r("Room 108",               "108",   _B, _FF, 1, col=4,  side="bottom"),
    "N_F_NOTICE": _fac("Notice Board",         _B, _FF, 1, col=5,  side="bottom", ftype="landmark"),
    "N_109":      _r("Analog & Digital Lab",   "109",   _B, _FF, 1, col=10, side="bottom"),
    "N_110":      _r("Computer Lab-3",         "110",   _B, _FF, 1, col=11, side="bottom"),
    "N_111":      _r("Room 111",               "111",   _B, _FF, 1, col=12, side="bottom"),
    # ── Second Floor ─────────────────────────────────────────────────────────
    "N_204":      _r("Digital Logic Design Lab","204",  _B, _SF, 2, col=2,  side="top"),
    "N_203":      _r("Linear & Digital App Lab","203",  _B, _SF, 2, col=3,  side="top"),
    "N_202":      _r("Tutorial Room",          "202",   _B, _SF, 2, col=4,  side="top"),
    "N_201":      _r("Tutorial Room",          "201",   _B, _SF, 2, col=5,  side="top"),
    "N_S_GENSWC": _fac("Gents Washroom",       _B, _SF, 2, col=5,  side="top", ftype="washroom"),  # FIX: missing node added
    "N_S_STAIR":  _fac("Staircase",            _B, _SF, 2, col=6,  side="top", ftype="stairs"),
    "N_S_LIFT":   _fac("Lift",                 _B, _SF, 2, col=7,  side="top", ftype="lift"),
    "N_S_STAIR2": _fac("Staircase (2)",        _B, _SF, 2, col=9,  side="top", ftype="stairs"),
    "N_S_GRLSWC": _fac("Girls Washroom",       _B, _SF, 2, col=10, side="top", ftype="washroom"),
    "N_216":      _r("Projects Lab",           "216",   _B, _SF, 2, col=11, side="top"),
    "N_215":      _r("Staff Room",             "215",   _B, _SF, 2, col=12, side="top"),
    "N_205":      _r("Room 205",               "205",   _B, _SF, 2, col=2,  side="bottom"),
    "N_206":      _r("Room 206",               "206",   _B, _SF, 2, col=3,  side="bottom"),
    "N_207":      _r("Electronic Circuit Analysis Lab","207", _B, _SF, 2, col=4, side="bottom"),
    "N_208":      _r("Electronic Devices & Circuit Lab","208", _B, _SF, 2, col=5, side="bottom"),
    "N_S_NOTICE": _fac("Notice Board",         _B, _SF, 2, col=6,  side="bottom", ftype="landmark"),
    "N_209":      _r("Room 209",               "209",   _B, _SF, 2, col=7,  side="bottom"),
    "N_210":      _r("Digital Electronics",    "210",   _B, _SF, 2, col=9,  side="bottom"),
    "N_211":      _r("Tutorial Room",          "211",   _B, _SF, 2, col=10, side="bottom"),
    "N_212":      _r("Room 212",               "212",   _B, _SF, 2, col=11, side="bottom"),
    "N_213":      _r("Tutorial Room",          "213",   _B, _SF, 2, col=12, side="bottom"),
    "N_214":      _r("Room 214",               "214",   _B, _SF, 2, col=13, side="bottom"),
    # ── Third Floor ──────────────────────────────────────────────────────────
    "N_305":      _r("Room 305",               "305",   _B, _TF, 3, col=2,  side="top"),
    "N_304":      _r("CSM Lab-2",              "304",   _B, _TF, 3, col=3,  side="top"),
    "N_303":      _r("CSM Lab-1",              "303",   _B, _TF, 3, col=4,  side="top"),
    "N_302":      _r("CSM HOD Cabin",          "302",   _B, _TF, 3, col=5,  side="top"),
    "N_T_GENSWC": _fac("Gents Washroom",       _B, _TF, 3, col=6,  side="top", ftype="washroom"),
    "N_T_STAIR":  _fac("Staircase",            _B, _TF, 3, col=7,  side="top", ftype="stairs"),
    "N_T_LIFT":   _fac("Lift",                 _B, _TF, 3, col=8,  side="top", ftype="lift"),
    "N_T_STAIR2": _fac("Staircase (2)",        _B, _TF, 3, col=9,  side="top", ftype="stairs"),
    "N_T_GRLSWC": _fac("Girls Washroom",       _B, _TF, 3, col=10, side="top", ftype="washroom"),
    "N_315":      _r("Drawing Hall-2",         "315",   _B, _TF, 3, col=11, side="top"),
    "N_314":      _r("Staff Room",             "314",   _B, _TF, 3, col=12, side="top"),
    "N_306":      _r("Room 306",               "306",   _B, _TF, 3, col=2,  side="bottom"),
    "N_307":      _r("Elements of Electrical", "307",   _B, _TF, 3, col=3,  side="bottom"),
    "N_T_NOTICE": _fac("Notice Board",         _B, _TF, 3, col=5,  side="bottom", ftype="landmark"),
    "N_308":      _r("Computer Lab-4",         "308",   _B, _TF, 3, col=9,  side="bottom"),
    "N_309":      _r("Common Room",            "309",   _B, _TF, 3, col=10, side="bottom"),
    "N_310":      _r("IT Workshop",            "310",   _B, _TF, 3, col=11, side="bottom"),
    "N_311":      _r("Room 311",               "311",   _B, _TF, 3, col=12, side="bottom"),
    "N_312":      _r("Tutorial Room",          "312",   _B, _TF, 3, col=13, side="bottom"),
})

# ══════════════════════════════════════════════════════════════════════════════
#  EAST BLOCK  (E_)
# ══════════════════════════════════════════════════════════════════════════════
_B, _BN = "E", "EAST BLOCK"
_SF2 = "Second Floor"

_NODES.update({
    # ── Ground Floor ─────────────────────────────────────────────────────────
    "E_ENTR":     _fac("East Block Entrance",  _BN, _GF, 0, col=4, side="bottom", ftype="entrance"),
    "E_G_STAIR":  _fac("Staircase",            _B,  _GF, 0, col=5, side="top",    ftype="stairs"),
    "E_G_LIFT":   _fac("Lift",                 _B,  _GF, 0, col=6, side="top",    ftype="lift"),
    # ── First Floor ──────────────────────────────────────────────────────────
    "E_101":      _r("Canteen",                "101",  _B, _FF, 1, col=2, side="top"),
    "E_102":      _r("Stationary",             "102",  _B, _FF, 1, col=3, side="top"),
    "E_103":      _r("Indoor Sports",          "103",  _B, _FF, 1, col=4, side="top"),
    "E_F_STAIR":  _fac("Staircase",            _B, _FF, 1, col=5, side="top",    ftype="stairs"),
    "E_F_LIFT":   _fac("Lift",                 _B, _FF, 1, col=6, side="top",    ftype="lift"),
    "E_F_ENTR":   _fac("First Floor Entrance", _B, _FF, 1, col=4, side="bottom", ftype="entrance"),
    # ── Second Floor ─────────────────────────────────────────────────────────
    "E_201":      _r("Management Cabin",       "201",  _B, _SF2, 2, col=3, side="top"),
    "E_S_STAIR":  _fac("Staircase",            _B, _SF2, 2, col=5, side="top",    ftype="stairs"),
    "E_S_LIFT":   _fac("Lift",                 _B, _SF2, 2, col=6, side="top",    ftype="lift"),
})

# ══════════════════════════════════════════════════════════════════════════════
#  WEST BLOCK  (W_)
# ══════════════════════════════════════════════════════════════════════════════
_B, _BN = "W", "WEST BLOCK"
_4F = "Fourth Floor"

_NODES.update({
    # ── Ground Floor ─────────────────────────────────────────────────────────
    "W_G01":      _r("Principal Office",       "G01",  _B, _GF, 0, col=2,  side="top"),
    "W_G02":      _r("Administrative",         "G02",  _B, _GF, 0, col=3,  side="top"),
    "W_G_GRLSWC": _fac("Girls Washroom",       _B, _GF, 0, col=4,  side="top",    ftype="washroom"),
    "W_G_STAIR":  _fac("Staircase",            _B, _GF, 0, col=5,  side="top",    ftype="stairs"),
    "W_G_GENSWC": _fac("Gents Washroom",       _B, _GF, 0, col=6,  side="top",    ftype="washroom"),
    "W_G05":      _r("Admission Section",      "G05",  _B, _GF, 0, col=7,  side="top"),
    "W_G06":      _r("Examination",            "G06",  _B, _GF, 0, col=8,  side="top"),
    "W_G_LIFT":   _fac("Lift",                 _B, _GF, 0, col=9,  side="top",    ftype="lift"),
    "W_ENTR":     _fac("West Block Entrance",  _BN, _GF, 0, col=5, side="bottom", ftype="entrance"),
    # ── First Floor ──────────────────────────────────────────────────────────
    "W_101":      _r("Board Room",             "101",  _B, _FF, 1, col=2,  side="top"),
    "W_F_STAIR":  _fac("Staircase",            _B, _FF, 1, col=4,  side="top",    ftype="stairs"),
    "W_102":      _r("Central Library",        "102",  _B, _FF, 1, col=6,  side="top"),
    "W_F_COMPCT": _r("Computer Center",        "",     _B, _FF, 1, col=7,  side="top"),
    "W_F_LIFT":   _fac("Lift",                 _B, _FF, 1, col=8,  side="top",    ftype="lift"),
    # ── Second Floor ─────────────────────────────────────────────────────────
    "W_S_LIB":    _r("Central Library",        "",     _B, _SF, 2, col=3,  side="top"),
    "W_S_STAIR":  _fac("Staircase",            _B, _SF, 2, col=5,  side="top",    ftype="stairs"),
    "W_S_LIFT":   _fac("Lift",                 _B, _SF, 2, col=8,  side="top",    ftype="lift"),
    # ── Third Floor ──────────────────────────────────────────────────────────
    "W_302":      _r("Computer Lab-5",         "302",  _B, _TF, 3, col=2,  side="top"),
    "W_303":      _r("Computer Lab-4",         "303",  _B, _TF, 3, col=3,  side="top"),
    "W_T_STAIR":  _fac("Staircase",            _B, _TF, 3, col=5,  side="top",    ftype="stairs"),
    "W_306":      _r("Admission Cell",         "306",  _B, _TF, 3, col=7,  side="top"),
    "W_307":      _r("Placement Office",       "307",  _B, _TF, 3, col=8,  side="top"),
    "W_301":      _r("NSE Cell",               "301",  _B, _TF, 3, col=2,  side="bottom"),
    "W_308":      _r("IQ AC Cell",             "308",  _B, _TF, 3, col=9,  side="top"),
    "W_T_LIFT":   _fac("Lift",                 _B, _TF, 3, col=10, side="top",    ftype="lift"),
    # ── Fourth Floor ─────────────────────────────────────────────────────────
    "W_404":      _r("Seminar Hall-1",         "404",  _B, _4F, 4, col=3,  side="top"),
    "W_402":      _r("Seminar Hall-2",         "402",  _B, _4F, 4, col=6,  side="top"),
    "W_4_STAIR":  _fac("Staircase",            _B, _4F, 4, col=5,  side="top",    ftype="stairs"),
    "W_4_LIFT":   _fac("Lift",                 _B, _4F, 4, col=8,  side="top",    ftype="lift"),
})

# ─────────────────────────────────────────────────────────────────────────────
#  CAMPUS OUTDOOR NODES
#  (type = "outdoor"; connected to block entrance nodes and each other)
# ─────────────────────────────────────────────────────────────────────────────

_CAMPUS_OUTDOOR: Dict[str, dict] = {
    "CAMPUS_SOUTH":   {"name": "South Block",          "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_NORTH":   {"name": "North Block",          "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_EAST":    {"name": "East Block",           "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_WEST":    {"name": "West Block",           "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_BENCH":   {"name": "Bench Park",           "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_CANTEEN": {"name": "Canteen & Stationary", "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_EWS":     {"name": "EWS Workshop",         "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_GATE1":   {"name": "Gate-1",               "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_GATE2":   {"name": "Gate-2 Entry",         "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_MOBIKES": {"name": "Mobikes Parking",      "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_OAA":     {"name": "Open Air Auditorium",  "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
    "CAMPUS_SPORTS":  {"name": "Sports Ground",        "type": "outdoor", "col": 0, "side": "corridor", "floor_num": -1, "floor": "Outdoor"},
}

# Campus edges: computed from real GPS coordinates loaded from College_Db.xlsx.
# Populated after _CAMPUS_GPS is filled; graph builder uses this list.
# _build_campus_edges_from_gps() re-generates this after GPS load.
_CAMPUS_EDGES: List[Tuple[str, str, float]] = []   # populated by _init_campus_edges() below

# ── Map: block entrance node → its campus outdoor node ──────────────────────
_ENTR_TO_CAMPUS = {
    "S_ENTR": "CAMPUS_SOUTH",
    "N_ENTR": "CAMPUS_NORTH",
    "E_ENTR": "CAMPUS_EAST",
    "W_ENTR": "CAMPUS_WEST",
}
_CAMPUS_TO_ENTR = {v: k for k, v in _ENTR_TO_CAMPUS.items()}

# ── Map campus display name → campus node key (for start_navigation lookup) ─
_CAMPUS_NAME_MAP: Dict[str, str] = {
    "SOUTH BLOCK":          "CAMPUS_SOUTH",
    "NORTH BLOCK":          "CAMPUS_NORTH",
    "EAST BLOCK":           "CAMPUS_EAST",
    "WEST BLOCK":           "CAMPUS_WEST",
    "BENCH PARK":           "CAMPUS_BENCH",
    "CANTEEN & STATIONARY": "CAMPUS_CANTEEN",
    "EWS WORKSHOP":         "CAMPUS_EWS",
    "GATE-1":               "CAMPUS_GATE1",
    "GATE-2 ENTRY":         "CAMPUS_GATE2",
    "MOBIKES PARKING":      "CAMPUS_MOBIKES",
    "OPEN AIR AUDITORIUM":  "CAMPUS_OAA",
    "SPORTS GROUND":        "CAMPUS_SPORTS",
}

# ─────────────────────────────────────────────────────────────────────────────
#  GPS LOADER  — reads real coordinates from College_Db.xlsx
#               Sheet: Main_buildingNames
#               Columns: S.No | Building Names | GPS location
#               GPS format: "lat, lon"  (e.g. "17.3537, 78.5097")
# ─────────────────────────────────────────────────────────────────────────────

def _load_gps_from_excel(explicit_path: str = "") -> Dict[str, Tuple[float, float]]:
    """
    Read GPS coordinates from College_Db.xlsx → Main_buildingNames sheet.
    Column layout: S.No | Building Names | GPS location
    GPS format: "lat, lon"  e.g. "17.354092, 78.508494"

    Args:
        explicit_path: pass the Excel file path directly (NavigationScreen
                       passes _find_excel() result here so no guessing needed).

    Returns dict: campus_node_key → (lat, lon).
    Falls back gracefully if file/sheet/column not found.
    """
    import os
    try:
        import openpyxl
    except ImportError:
        print("[GPS] openpyxl not available – using fallback coords")
        return {}

    path = explicit_path.strip() if explicit_path else ""

    if not path or not os.path.isfile(path):
        # Fallback search
        try:
            here = os.path.dirname(os.path.abspath(__file__))
        except Exception:
            here = os.getcwd()
        cwd = os.getcwd()
        fnames = ["College_Db.xlsx", "college_db.xlsx", "College_DB.xlsx"]
        search_dirs: list = []
        for base in [here, cwd]:
            search_dirs += [
                base,
                os.path.join(base, "data"),
                os.path.join(base, "assets"),
                os.path.join(base, ".."),
                os.path.join(base, "..", "data"),
            ]
        seen: set = set()
        for d in search_dirs:
            d = os.path.normpath(d)
            if d in seen:
                continue
            seen.add(d)
            for fn in fnames:
                p = os.path.join(d, fn)
                if os.path.isfile(p):
                    path = p
                    break
            if path:
                break

    if not path or not os.path.isfile(path):
        print("[GPS] College_Db.xlsx not found – using fallback GPS coords")
        return {}

    result: Dict[str, Tuple[float, float]] = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Main_buildingNames" not in wb.sheetnames:
            print("[GPS] Main_buildingNames sheet missing")
            wb.close()
            return {}

        ws = wb["Main_buildingNames"]
        rows = list(ws.iter_rows(values_only=True))

        # Row 0 = headers (S.No | Building Names | GPS location)
        # Row 1+ = data
        gps_col_idx = None
        name_col_idx = 1   # default
        if rows:
            hdr = [str(c).strip().lower() if c else "" for c in rows[0]]
            for idx, h in enumerate(hdr):
                if "gps" in h or "location" in h or "lat" in h or "coord" in h:
                    gps_col_idx = idx
                if "building" in h or "name" in h:
                    name_col_idx = idx

        if gps_col_idx is None:
            # Try column index 2 (third column) as default
            gps_col_idx = 2

        for row in rows[1:]:
            if not row or not any(row):
                continue
            bname = row[name_col_idx] if len(row) > name_col_idx else None
            gps   = row[gps_col_idx]  if len(row) > gps_col_idx  else None

            if not bname or not gps:
                continue

            bname_str = str(bname).strip()
            gps_str   = str(gps).strip()

            # Parse "lat, lon"
            try:
                parts = gps_str.replace(";", ",").split(",")
                if len(parts) >= 2:
                    lat = float(parts[0].strip())
                    lon = float(parts[1].strip())
                else:
                    print(f"[GPS] Cannot parse GPS for {bname_str!r}: {gps_str!r}")
                    continue
            except ValueError:
                print(f"[GPS] Invalid GPS value for {bname_str!r}: {gps_str!r}")
                continue

            # Map building name → campus node key
            campus_key = _CAMPUS_NAME_MAP.get(bname_str.upper())
            if campus_key:
                result[campus_key] = (lat, lon)
                print(f"[GPS] {bname_str:<24} → {campus_key}  ({lat:.6f}, {lon:.6f})")
            else:
                print(f"[GPS] No mapping for {bname_str!r}")

        wb.close()
        print(f"[GPS] Loaded {len(result)} GPS coordinates from {os.path.basename(path)}")
    except Exception as exc:
        print(f"[GPS] Excel read error: {exc}")

    return result


# ── Fallback hardcoded GPS (used only if Excel is unavailable) ────────────────
_CAMPUS_GPS_FALLBACK: Dict[str, Tuple[float, float]] = {
    "CAMPUS_SOUTH":   (17.354092, 78.508494),
    "CAMPUS_NORTH":   (17.354543, 78.508580),
    "CAMPUS_EAST":    (17.353730, 78.509749),
    "CAMPUS_WEST":    (17.354358, 78.508198),
    "CAMPUS_BENCH":   (17.354530, 78.509077),
    "CAMPUS_CANTEEN": (17.353820, 78.509749),  # Canteen inside East Block (slightly offset from CAMPUS_EAST)
    "CAMPUS_EWS":     (17.355038, 78.508888),
    "CAMPUS_GATE1":   (17.352993, 78.509112),
    "CAMPUS_GATE2":   (17.353836, 78.509222),
    "CAMPUS_MOBIKES": (17.353663, 78.509616),
    "CAMPUS_OAA":     (17.353775, 78.509022),
    "CAMPUS_SPORTS":  (17.354430, 78.509703),
}

# Load from Excel; merge with fallback so any missing entries are covered
_gps_from_excel = _load_gps_from_excel()
_CAMPUS_GPS: Dict[str, Tuple[float, float]] = {**_CAMPUS_GPS_FALLBACK, **_gps_from_excel}

# Pre-compute Haversine distances between every GPS pair for use by _build_graph.
# These are recalculated after reload_campus_gps() is called from the screen.
def _init_campus_edges() -> List[Tuple[str, str, float]]:
    """Compute Haversine distances between every GPS pair.
    Inlines the math so this can be called before _haversine is defined."""
    _R = 6_371_000.0
    edges = []
    keys = list(_CAMPUS_GPS.keys())
    for i, a in enumerate(keys):
        for b in keys[i+1:]:
            lat1, lon1 = _CAMPUS_GPS[a]
            lat2, lon2 = _CAMPUS_GPS[b]
            p1, p2 = _math.radians(lat1), _math.radians(lat2)
            dp = _math.radians(lat2 - lat1)
            dl = _math.radians(lon2 - lon1)
            x = _math.sin(dp/2)**2 + _math.cos(p1)*_math.cos(p2)*_math.sin(dl/2)**2
            dist = _R * 2 * _math.atan2(_math.sqrt(x), _math.sqrt(1 - x))
            edges.append((a, b, dist))
    return edges


# Populate _CAMPUS_EDGES at module load (before _build_graph is called)
_CAMPUS_EDGES = _init_campus_edges()


def _build_campus_edges_from_gps() -> List[Tuple[str, str, float]]:
    """Recompute Haversine edges after GPS coords have been reloaded."""
    return _init_campus_edges()


def reload_campus_gps(excel_path: str = "") -> bool:
    """
    Public function called by NavigationScreen after it locates College_Db.xlsx.
    Reads GPS from the sheet, updates _CAMPUS_GPS, and re-weights the campus
    edges in _GRAPH with real Haversine distances.

    Returns True if at least one GPS coordinate was loaded.
    """
    global _CAMPUS_GPS, _GRAPH
    loaded = _load_gps_from_excel(excel_path)
    if not loaded:
        print("[GPS] reload_campus_gps: no data loaded")
        return False

    # Merge: Excel values override fallbacks
    _CAMPUS_GPS = {**_CAMPUS_GPS_FALLBACK, **loaded}

    # Re-weight all campus-to-campus edges in the live graph
    new_edges = _build_campus_edges_from_gps()
    for a, b, dist in new_edges:
        # Use 'in _GRAPH' instead of _GRAPH.has_node() — works for both
        # networkx.Graph and any dict-like graph object
        if a in _GRAPH and b in _GRAPH:
            if _GRAPH.has_edge(a, b):
                _GRAPH[a][b]["weight"] = dist
            else:
                _GRAPH.add_edge(a, b, weight=dist)

    print(f"[GPS] reload_campus_gps: {len(loaded)} locations, "
          f"{len(new_edges)} campus edges updated with real distances")
    return True


# Public accessors for NavigationScreen (display GPS values in UI)
def get_campus_gps() -> Dict[str, Tuple[float, float]]:
    """Return the live GPS dict (Excel-loaded + fallback merged)."""
    return dict(_CAMPUS_GPS)


def get_campus_gps_display() -> Dict[str, str]:
    """Return human-readable 'lat, lon' strings for every campus location."""
    return {k: f"{v[0]:.6f}, {v[1]:.6f}" for k, v in _CAMPUS_GPS.items()}


# ─────────────────────────────────────────────────────────────────────────────
#  LIFT & STAIR SPINE  (per block, per floor)
# ─────────────────────────────────────────────────────────────────────────────
_LIFT_NODE: Dict[str, Dict[int, str]] = {
    "S": {0: "S_G_LIFT",  1: "S_F_LIFT",  2: "S_S_LIFT",  3: "S_T_LIFT"},
    "N": {0: "N_G_LIFT",  1: "N_F_LIFT",  2: "N_S_LIFT",  3: "N_T_LIFT"},
    "E": {0: "E_G_LIFT",  1: "E_F_LIFT",  2: "E_S_LIFT"},
    "W": {0: "W_G_LIFT",  1: "W_F_LIFT",  2: "W_S_LIFT",  3: "W_T_LIFT",  4: "W_4_LIFT"},
}
_STAIR_NODE: Dict[str, Dict[int, str]] = {
    "S": {0: "S_G_STAIR",  1: "S_F_STAIR",  2: "S_S_STAIR",  3: "S_T_STAIR"},
    "N": {0: "N_G_STAIR",  1: "N_F_STAIR",  2: "N_S_STAIR",  3: "N_T_STAIR"},
    "E": {0: "E_G_STAIR",  1: "E_F_STAIR",  2: "E_S_STAIR"},
    "W": {0: "W_G_STAIR",  1: "W_F_STAIR",  2: "W_S_STAIR",  3: "W_T_STAIR",  4: "W_4_STAIR"},
}
_STAIR2_NODE: Dict[str, Dict[int, str]] = {
    "S": {1: "S_F_STAIR2", 2: "S_S_STAIR2", 3: "S_T_STAIR2"},
    "N": {1: "N_F_STAIR2", 2: "N_S_STAIR2", 3: "N_T_STAIR2"},
}


# ─────────────────────────────────────────────────────────────────────────────
#  BUILD UNIFIED GRAPH
# ─────────────────────────────────────────────────────────────────────────────

def _cor_id(blk: str, fl: int, col: int) -> str:
    return f"{blk}_{_FC[fl]}C{col}"


def _build_graph():
    G = nx.Graph()

    # ── 1. Indoor nodes ───────────────────────────────────────────────────────
    for nid, attrs in _NODES.items():
        G.add_node(nid, **attrs)

    # ── 2. Corridor spine per block ───────────────────────────────────────────
    for blk, fl_ranges in _COR_RANGES.items():
        for fl, (cmin, cmax) in fl_ranges.items():
            fl_name = _FLOOR_NAME[fl]
            # Create corridor waypoints
            for col in range(cmin, cmax + 1):
                cid = _cor_id(blk, fl, col)
                G.add_node(cid, name=f"{fl_name} Corridor",
                           room_num="", block=_BLOCK_LONG.get(blk, blk),
                           floor=fl_name, floor_num=fl,
                           type="corridor", col=col, side="corridor")
            # Connect corridor waypoints in a line
            for col in range(cmin, cmax):
                G.add_edge(_cor_id(blk, fl, col),
                           _cor_id(blk, fl, col + 1),
                           weight=_W_COR)

    # ── 3. Connect rooms/facilities to their corridor waypoint ────────────────
    for nid, attrs in _NODES.items():
        blk_char = nid[0]
        if blk_char not in _COR_RANGES:
            continue
        fl  = attrs["floor_num"]
        col = attrs["col"]
        if fl not in _COR_RANGES[blk_char]:
            continue
        cmin, cmax = _COR_RANGES[blk_char][fl]
        cor_col = max(cmin, min(cmax, col))
        G.add_edge(nid, _cor_id(blk_char, fl, cor_col), weight=_W_COR)

    # ── 4. Inter-floor: lifts ─────────────────────────────────────────────────
    for blk, fl_map in _LIFT_NODE.items():
        floors = sorted(fl_map.keys())
        for a, b in zip(floors, floors[1:]):
            if fl_map[a] in G and fl_map[b] in G:
                G.add_edge(fl_map[a], fl_map[b], weight=_W_LIFT)

    # ── 5. Inter-floor: main staircases ───────────────────────────────────────
    for blk, fl_map in _STAIR_NODE.items():
        floors = sorted(fl_map.keys())
        for a, b in zip(floors, floors[1:]):
            if fl_map[a] in G and fl_map[b] in G:
                G.add_edge(fl_map[a], fl_map[b], weight=_W_STAIR)

    # ── 6. Inter-floor: secondary staircases ─────────────────────────────────
    for blk, fl_map in _STAIR2_NODE.items():
        floors = sorted(fl_map.keys())
        for a, b in zip(floors, floors[1:]):
            if fl_map[a] in G and fl_map[b] in G:
                G.add_edge(fl_map[a], fl_map[b], weight=_W_STAIR)

    # ── 7. Campus outdoor nodes ───────────────────────────────────────────────
    for cid, attrs in _CAMPUS_OUTDOOR.items():
        G.add_node(cid, **attrs)

    for a, b, dist in _CAMPUS_EDGES:
        if a in G and b in G:
            G.add_edge(a, b, weight=dist)

    # ── 8. Connect block entrances to campus outdoor nodes ────────────────────
    for entr_id, campus_id in _ENTR_TO_CAMPUS.items():
        if entr_id in G and campus_id in G:
            G.add_edge(entr_id, campus_id, weight=_W_DOOR)

    return G


_GRAPH = _build_graph()


# ─────────────────────────────────────────────────────────────────────────────
#  ROOM-DISPLAY → NODE-ID LOOKUP
# ─────────────────────────────────────────────────────────────────────────────

# Map: "NORTH BLOCK" → "N", etc.
_BLK_PREFIX = {
    "SOUTH BLOCK": "S", "NORTH BLOCK": "N",
    "EAST BLOCK":  "E", "WEST BLOCK":  "W",
}
 

def _extract_room_key(display: str) -> str:
    """
    Resolve a display string to a graph node ID.

    Formats accepted
    ----------------
      "NORTH BLOCK|Computer Lab-2 - 107"   → "N_107"
      "SOUTH BLOCK|IT HOD - 302"            → "S_302"
      "SOUTH BLOCK|Girls Washroom"          → fuzzy match in S_ nodes
      Legacy (no block prefix):
        "CSE-A 2nd yr - G13"               → "S_G13"  (falls back to South)
        "305"                              → searches all nodes
    """
    blk_pfx = ""
    room_part = display.strip()

    # ── Split block prefix if present ────────────────────────────────────────
    if "|" in display:
        blk_key, room_part = display.split("|", 1)
        blk_pfx = _BLK_PREFIX.get(blk_key.strip().upper(), "")
        room_part = room_part.strip()

    # ── Try "Name - RoomNo" pattern ───────────────────────────────────────────
    if " - " in room_part:
        candidate = room_part.rsplit(" - ", 1)[-1].strip()
        candidate_norm = re.sub(r"[,\s]+", "", candidate)

        # With block prefix
        if blk_pfx:
            nid = f"{blk_pfx}_{candidate_norm}"
            if nid in _NODES:
                return nid

        # Search all blocks
        for pfx in (["S", "N", "E", "W"] if not blk_pfx else [blk_pfx]):
            nid = f"{pfx}_{candidate_norm}"
            if nid in _NODES:
                return nid

    # ── Try full display as raw node ID ───────────────────────────────────────
    display_norm = re.sub(r"[,\s]+", "", room_part.strip())
    for pfx in (["S", "N", "E", "W"] if not blk_pfx else [blk_pfx]):
        nid = f"{pfx}_{display_norm}"
        if nid in _NODES:
            return nid
    if display_norm in _NODES:
        return display_norm

    # ── Fuzzy: name substring match (restrict to block if known) ──────────────
    low = room_part.strip().lower()
    candidates = {k: v for k, v in _NODES.items()
                  if (not blk_pfx or k.startswith(blk_pfx + "_"))}
    for nid, attrs in candidates.items():
        name_low = attrs["name"].lower()
        if name_low in low or low in name_low:
            return nid

    return ""


# ─────────────────────────────────────────────────────────────────────────────
#  NODE-INFO HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _node_info(nid: str) -> dict:
    if nid in _NODES:
        return _NODES[nid]
    if nid in _CAMPUS_OUTDOOR:
        return _CAMPUS_OUTDOOR[nid]
    # Safely access graph node attributes — _GRAPH.nodes can behave as a
    # method in some networkx versions, so guard against that.
    try:
        node_data = _GRAPH.nodes[nid]
        return dict(node_data)
    except (TypeError, KeyError):
        pass
    try:
        # Fallback: iterate nodes to find the matching one
        # NOTE: _SimpleGraph.nodes is a plain dict property — use .items(), not callable(data=True)
        for n, attrs in _GRAPH.nodes.items():
            if n == nid:
                return dict(attrs)
    except Exception:
        pass
    return {}


def _label(nid: str) -> str:
    info = _node_info(nid)
    name = info.get("name", nid)
    rnum = info.get("room_num", "")
    if rnum and rnum not in ("-", ""):
        return f"{name} ({rnum})"
    return name


def _block_of(nid: str) -> str:
    """Return block short prefix ('S','N','E','W') or '' for campus nodes."""
    if nid in _NODES:
        return _NODES[nid].get("block", "")
    return ""


# ─────────────────────────────────────────────────────────────────────────────
#  DIRECTION & SIDE HELPERS  (unchanged logic, generalised)
# ─────────────────────────────────────────────────────────────────────────────

def _direction(from_col: int, to_col: int) -> str:
    if to_col > from_col: return "RIGHT"
    if to_col < from_col: return "LEFT"
    return "STRAIGHT"


def _room_side_hint(room_id: str, walk_dir: str) -> str:
    side = _node_info(room_id).get("side", "top")
    if walk_dir == "RIGHT": return "LEFT" if side == "top" else "RIGHT"
    if walk_dir == "LEFT":  return "RIGHT" if side == "top" else "LEFT"
    return "LEFT" if side == "top" else "RIGHT"


def _invert_if_top_side(side: str, walk_dir: str) -> str:
    if side == "top":
        if walk_dir == "LEFT":  return "RIGHT"
        if walk_dir == "RIGHT": return "LEFT"
    return walk_dir


def _get_travel_direction(path: List[str], idx: int) -> str:
    curr_col = _node_info(path[idx]).get("col", 0)
    for k in range(idx - 1, -1, -1):
        pc = _node_info(path[k]).get("col", curr_col)
        if pc != curr_col:
            return _direction(pc, curr_col)
    for k in range(idx + 1, len(path)):
        fc = _node_info(path[k]).get("col", curr_col)
        if fc != curr_col:
            return _direction(curr_col, fc)
    return "STRAIGHT"


def _landmarks_between(blk: str, fl: int, c_start: int, c_end: int,
                        exclude: set) -> List[str]:
    lo, hi = min(c_start, c_end), max(c_start, c_end)
    going_right = c_end > c_start
    items: List[Tuple[int, str]] = []
    for nid, attrs in _NODES.items():
        if nid in exclude: continue
        if attrs.get("block", "")[:1] != blk and attrs.get("block", "") != _BLOCK_LONG.get(blk, blk): continue
        if attrs.get("floor_num", -1) != fl: continue
        c = attrs.get("col", -1)
        if lo < c < hi and attrs.get("type", "") in ("room", "washroom", "stairs", "lift", "entrance"):
            items.append((c, attrs.get("name", nid)))
    items.sort(key=lambda x: x[0], reverse=not going_right)
    return [name for _, name in items]


# ─────────────────────────────────────────────────────────────────────────────
#  INSTRUCTION GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def _outdoor_segment_distance(path: List[str], start_idx: int) -> float:
    """
    Sum edge weights for the outdoor walking segment starting at start_idx.
    Stops BEFORE the destination entrance so it only counts the actual walk.
    """
    total = 0.0
    first = True
    for k in range(start_idx, len(path) - 1):
        nt = _node_info(path[k]).get("type", "")
        if not first and nt == "entrance":
            break          # arrived at destination building – don't add its door
        if nt not in ("outdoor", "entrance"):
            break
        total += _GRAPH.edges[path[k], path[k + 1]].get("weight", 0)
        first = False
    return total


def _generate_instructions(path: List[str]) -> List[str]:
    if not path or len(path) < 2:
        return ["You are already at your destination."]

    steps: List[str] = []
    all_ids = set(path)
    i = 0

    while i < len(path) - 1:
        curr      = path[i]
        curr_info = _node_info(curr)
        curr_type = curr_info.get("type", "corridor")
        curr_fl   = curr_info.get("floor_num", 0)
        curr_col  = curr_info.get("col", 0)
        curr_blk  = curr[0] if len(curr) > 0 and curr[0] in "SNEW" else ""

        next_id   = path[i + 1]
        next_info = _node_info(next_id)
        next_type = next_info.get("type", "corridor")

        # ── (A) Room/entrance → corridor ─────────────────────────────────────
        if curr_type in ("room", "entrance") and next_type == "corridor":
            walk_dir = "STRAIGHT"
            for k in range(i + 1, len(path)):
                kc = _node_info(path[k]).get("col", curr_col)
                if kc != curr_col:
                    walk_dir = _direction(curr_col, kc)
                    break
            walk_dir = _invert_if_top_side(curr_info.get("side", "bottom"), walk_dir)
            steps.append(
                f"You are at {_label(curr)}. "
                f"Exit and turn {walk_dir} along the corridor."
            )
            i += 1; continue

        # ── (B) Corridor → corridor ───────────────────────────────────────────
        if curr_type == "corridor" and next_type == "corridor":
            run = [curr]
            j = i + 1
            while j < len(path) and _node_info(path[j]).get("type", "corridor") == "corridor":
                run.append(path[j])
                j += 1
            end_col  = _node_info(run[-1]).get("col", curr_col)
            walk_dir = _direction(curr_col, end_col)
            lm = _landmarks_between(curr_blk, curr_fl, curr_col, end_col, all_ids)
            if lm:
                steps.append(f"Walk {walk_dir} along the corridor, passing {', '.join(lm[:3])}.")
            else:
                steps.append(f"Walk {walk_dir} along the corridor.")
            i = j - 1; continue

        # ── (C) Corridor → room ───────────────────────────────────────────────
        if curr_type == "corridor" and next_type == "room":
            walk_dir  = _get_travel_direction(path, i)
            side_hint = _room_side_hint(next_id, walk_dir)
            steps.append(
                f"Room {_label(next_id)} is on your {side_hint}. "
                f"Enter your destination."
            )
            i += 1; continue

        # ── (D) Corridor → facility (lift/stairs/washroom/entrance/landmark) ──
        if curr_type == "corridor" and next_type in ("lift", "stairs", "washroom", "entrance", "landmark"):
            walk_dir  = _get_travel_direction(path, i)
            side_hint = _room_side_hint(next_id, walk_dir)
            steps.append(f"The {next_info.get('name', next_id)} is on your {side_hint}.")
            i += 1; continue

        # ── (E) Lift → lift  (floor change) ──────────────────────────────────
        if curr_type == "lift" and next_type == "lift":
            steps.append(
                f"Take the Lift from {curr_info.get('floor', '')} to {next_info.get('floor', '')}."
            )
            i += 1; continue

        # ── (F) Stairs → stairs  (floor change) ──────────────────────────────
        if curr_type == "stairs" and next_type == "stairs":
            steps.append(
                f"Take the Staircase from {curr_info.get('floor', '')} to {next_info.get('floor', '')}."
            )
            i += 1; continue

        # ── (G) Lift/stairs → corridor  (landed on new floor) ────────────────
        if curr_type in ("lift", "stairs") and next_type == "corridor":
            floor_name = next_info.get("floor", "")
            j = i + 1
            while j < len(path) and _node_info(path[j]).get("type", "corridor") == "corridor":
                j += 1
            if j < len(path):
                dest_col = _node_info(path[j]).get("col", curr_col)
                walk_dir = _direction(curr_col, dest_col)
                if walk_dir == "STRAIGHT" and j > i + 1:
                    cs = _node_info(path[i + 1]).get("col", curr_col)
                    ce = _node_info(path[j - 1]).get("col", curr_col)
                    walk_dir = _direction(cs, ce) if cs != ce else "STRAIGHT"
            else:
                walk_dir = "STRAIGHT"
            walk_dir = _invert_if_top_side(curr_info.get("side", "bottom"), walk_dir)
            facility = "Lift" if curr_type == "lift" else "Staircase"
            steps.append(
                f"Exit the {facility} on the {floor_name}. "
                f"Turn {walk_dir} along the corridor."
            )
            i += 1; continue

        # ── (H) Room → lift/stairs directly ──────────────────────────────────
        if curr_type == "room" and next_type in ("lift", "stairs"):
            steps.append(
                f"Walk from {_label(curr)} to the {next_info.get('name', next_id)} on this floor."
            )
            i += 1; continue

        # ── (I) Indoor entrance → outdoor campus node ─────────────────────────
        if curr_type == "entrance" and next_type == "outdoor":
            blk_raw  = curr_info.get("block", curr_info.get("name", "building"))
            # Resolve to proper title-case block name
            blk_name = _BLOCK_LONG.get(blk_raw[:1].upper(), blk_raw) if len(blk_raw) == 1 \
                       else blk_raw.title()
            total_m  = _outdoor_segment_distance(path, i)
            # Find destination name at end of outdoor segment
            j = i + 1
            while j < len(path) - 1 and _node_info(path[j]).get("type", "") == "outdoor":
                j += 1
            dest_raw  = _node_info(path[j]).get("name", "destination")
            dest_name = dest_raw.title() if dest_raw.isupper() else dest_raw
            steps.append(
                f"🚶 Exit {blk_name}.  Walk ~{total_m:.0f} m to {dest_name}."
            )
            i += 1; continue

        # ── (J) Outdoor → outdoor  (waypoint hop) ────────────────────────────
        if curr_type == "outdoor" and next_type == "outdoor":
            # If the node AFTER the next one is a building entrance, skip this
            # intermediate hop — step I already shows the total distance and
            # step K will announce entering the building.
            nxt_next_type = (_node_info(path[i + 2]).get("type", "")
                             if i + 2 < len(path) else "")
            if nxt_next_type != "entrance":
                dist     = _GRAPH.edges[curr, next_id].get("weight", 0)
                wpt_name = next_info.get("name", next_id)
                steps.append(f"Walk ~{dist:.0f} m towards {wpt_name}.")
            i += 1; continue

        # ── (K) Outdoor → building entrance  ─────────────────────────────────
        if curr_type == "outdoor" and next_type == "entrance":
            blk_raw  = next_info.get("block", next_info.get("name", "building"))
            blk_name = _BLOCK_LONG.get(blk_raw[:1].upper(), blk_raw) if len(blk_raw) == 1 \
                       else blk_raw.title()
            steps.append(f"🏛️  You have reached {blk_name}.  Enter the building.")
            i += 1; continue

        # ── (L) Entrance → corridor (entered building at ground floor) ────────
        if curr_type == "entrance" and next_type == "corridor":
            walk_dir = "STRAIGHT"
            for k in range(i + 1, len(path)):
                kc = _node_info(path[k]).get("col", curr_col)
                if kc != curr_col:
                    walk_dir = _direction(curr_col, kc)
                    break
            walk_dir = _invert_if_top_side(curr_info.get("side", "bottom"), walk_dir)
            steps.append(
                f"You are at {curr_info.get('name', 'Entrance')}. "
                f"Go inside and turn {walk_dir} along the corridor."
            )
            i += 1; continue

        # ── Fallback ──────────────────────────────────────────────────────────
        i += 1

    # Arrival
    dst = path[-1]
    blk_raw   = _node_info(dst).get("block", "")
    blk_label = _BLOCK_LONG.get(blk_raw[:1].upper(), blk_raw) if (blk_raw and len(blk_raw) == 1) \
                else blk_raw
    blk_suffix = f" [{blk_label}]" if blk_label and blk_label not in ("", "Outdoor") else ""
    steps.append(f"🏁  You have arrived at {_label(dst)}{blk_suffix}.")
    return steps


# ─────────────────────────────────────────────────────────────────────────────
#  GPS / BEARING MATHS
# ─────────────────────────────────────────────────────────────────────────────
_EARTH_R = 6_371_000.0


def _haversine(lat1, lon1, lat2, lon2) -> float:
    phi1, phi2 = _math.radians(lat1), _math.radians(lat2)
    dphi       = _math.radians(lat2 - lat1)
    dl         = _math.radians(lon2 - lon1)
    a = _math.sin(dphi/2)**2 + _math.cos(phi1)*_math.cos(phi2)*_math.sin(dl/2)**2
    return _EARTH_R * 2 * _math.atan2(_math.sqrt(a), _math.sqrt(1 - a))


def _bearing(lat1, lon1, lat2, lon2) -> float:
    phi1, phi2 = _math.radians(lat1), _math.radians(lat2)
    dl = _math.radians(lon2 - lon1)
    x  = _math.sin(dl) * _math.cos(phi2)
    y  = _math.cos(phi1)*_math.sin(phi2) - _math.sin(phi1)*_math.cos(phi2)*_math.cos(dl)
    return (_math.degrees(_math.atan2(x, y)) + 360) % 360


def _arrow_offset(bearing_deg, heading_deg, cx, cy, length=80):
    a  = _math.radians(bearing_deg - heading_deg)
    dx = _math.sin(a) * length
    dy = -_math.cos(a) * length
    return (int(cx - dx*.5), int(cy - dy*.5)), (int(cx + dx*.5), int(cy + dy*.5))


def _bearing_to_compass(bearing_deg: float) -> str:
    """Convert 0-360 bearing to 16-point compass label."""
    dirs = ["N","NNE","NE","ENE","E","ESE","SE","SSE",
            "S","SSW","SW","WSW","W","WNW","NW","NNW"]
    return dirs[int((bearing_deg + 11.25) / 22.5) % 16]


# ─────────────────────────────────────────────────────────────────────────────
#  NAVIGATION CONTROLLER  (public class)
# ─────────────────────────────────────────────────────────────────────────────

class NavigationController:
    """
    Manages one navigation session.  Instantiated fresh on each on_enter().

    Key fields
    ----------
    is_navigating        bool
    is_inter_building    bool   True for GPS/AR campus-level navigation
    is_cross_block       bool   True when path crosses indoor blocks
    current_instruction_idx  int  (writable)
    """

    ARRIVAL_RADIUS_M = 15.0

    def __init__(self):
        self.is_navigating           = False
        self.is_inter_building       = False
        self.is_cross_block          = False
        self.current_instruction_idx = 0
        self._instructions: List[str] = []
        self._path:         List[str] = []
        self._src_node  = ""
        self._dst_node  = ""

        # GPS / compass
        self.gps_lat:          Optional[float] = None
        self.gps_lon:          Optional[float] = None
        self.compass_heading:  float           = 0.0
        self.live_distance_m:  Optional[float] = None
        self.live_bearing_deg: Optional[float] = None
        self._dst_lat: Optional[float] = None
        self._dst_lon: Optional[float] = None
        self._src_lat: Optional[float] = None
        self._src_lon: Optional[float] = None
        self._src_campus_key: str = ""
        self._dst_campus_key: str = ""
        self.src_gps_display: str = ""
        self.dst_gps_display: str = ""
        self.static_distance_m: Optional[float] = None
        # ar_overlay_active: True when AR compass/distance overlay should draw
        # even for cross-block STEPS navigation (one side is indoor, other is campus)
        self.ar_overlay_active: bool = False
        self.on_arrive_callback = None

    # ── GPS / compass feed-in ─────────────────────────────────────────────────

    def on_gps_update(self, lat: float, lon: float) -> None:
        self.gps_lat = lat
        self.gps_lon = lon
        if (self._dst_lat is not None and self._dst_lon is not None
                and self.is_navigating
                and (self.is_inter_building or self.ar_overlay_active)):
            self.live_distance_m  = _haversine(lat, lon, self._dst_lat, self._dst_lon)
            self.live_bearing_deg = _bearing(lat, lon, self._dst_lat, self._dst_lon)
            if self.live_distance_m <= self.ARRIVAL_RADIUS_M:
                if self.on_arrive_callback:
                    self.on_arrive_callback()

    def on_compass_update(self, heading: float) -> None:
        self.compass_heading = heading

    def enable_ar_overlay(self, campus_key: str) -> bool:
        """
        Activate AR compass/distance overlay for a mixed navigation session.
        Steps path is already set; this adds the GPS destination for the
        outdoor end so render_ar_frame can draw the directional arrow.

        Called from NavigationScreen Branch B after start_navigation() succeeds.
        Returns True on success.
        """
        gps = _CAMPUS_GPS.get(campus_key)
        if gps is None:
            print(f"[AR] enable_ar_overlay: no GPS for {campus_key!r}")
            return False

        self._dst_lat, self._dst_lon = gps
        self._dst_campus_key = campus_key
        self.dst_gps_display = f"{gps[0]:.6f}, {gps[1]:.6f}"

        # Try to get a source GPS reference for initial bearing/distance
        src_gps = _CAMPUS_GPS.get(self._src_campus_key) if self._src_campus_key else None
        if src_gps:
            self._src_lat, self._src_lon = src_gps
            self.src_gps_display   = f"{src_gps[0]:.6f}, {src_gps[1]:.6f}"
            self.static_distance_m = _haversine(src_gps[0], src_gps[1], gps[0], gps[1])
            self.live_distance_m   = self.static_distance_m
            self.live_bearing_deg  = _bearing(src_gps[0], src_gps[1], gps[0], gps[1])
        else:
            self.src_gps_display   = "Indoor start"
            self.static_distance_m = None
            self.live_distance_m   = None
            self.live_bearing_deg  = None

        self.ar_overlay_active = True
        print(f"[AR] overlay enabled → {campus_key}  GPS={gps[0]:.6f},{gps[1]:.6f}  "
              f"dist={self.static_distance_m}")
        return True

    # ── Public: start ─────────────────────────────────────────────────────────

    def start_navigation(self, src_display: str, dst_display: str) -> Tuple[bool, str]:
        """
        Resolve src/dst → graph path → instructions + GPS distance.

        AR / distance rules
        ───────────────────
        • SAME block rooms (e.g. E_101 → E_102)  → Steps only, NO AR
        • DIFFERENT blocks (e.g. E_xxx → W_yyy)  → AR compass + Steps
          GPS used = block-level campus pin for each side
        • Both campus/main-building locations     → full AR, no step list
        • One campus + one indoor room            → AR compass + Steps

        GPS distance is always computed so the AR panel never shows
        "Calculating…" — uses Haversine between block GPS pins, or the
        graph path weight when both pins are identical (same block).
        """
        try:
            return self._start_navigation_impl(src_display, dst_display)
        except Exception as exc:
            import traceback; traceback.print_exc()
            return False, f"Internal navigation error: {exc}"

    def _start_navigation_impl(self, src_display: str, dst_display: str) -> Tuple[bool, str]:

        # ── Helper: map display string → campus-node key (None = indoor room) ──
        def _to_campus_key(raw: str) -> Optional[str]:
            upper = raw.strip().upper()
            if upper in _CAMPUS_NAME_MAP:
                return _CAMPUS_NAME_MAP[upper]
            if "|" in upper:
                blk_part, room_part = upper.split("|", 1)
                blk_part  = blk_part.strip()
                room_part = room_part.strip()
                # "East block main Building", "South Block main Building", etc.
                if (room_part.upper().endswith("MAIN BUILDING")
                        or "MAIN BUILDING" in room_part.upper()):
                    return _CAMPUS_NAME_MAP.get(blk_part)
            return None

        # ── Helper: campus GPS pin for any node ──────────────────────────────
        def _node_campus_key(node_id: str) -> str:
            """Return the _CAMPUS_GPS key that represents this node's location."""
            if node_id in _CAMPUS_GPS:            # already a campus node
                return node_id
            if node_id.startswith("E_"):  return "CAMPUS_EAST"
            if node_id.startswith("S_"):  return "CAMPUS_SOUTH"
            if node_id.startswith("N_"):  return "CAMPUS_NORTH"
            if node_id.startswith("W_"):  return "CAMPUS_WEST"
            return ""

        def _node_gps(node_id: str) -> Optional[Tuple[float, float]]:
            key = _node_campus_key(node_id)
            return _CAMPUS_GPS.get(key) if key else None

        # ── Classify each endpoint ───────────────────────────────────────────
        src_campus = _to_campus_key(src_display)   # None = indoor room
        dst_campus = _to_campus_key(dst_display)
        is_src_campus = src_campus is not None
        is_dst_campus = dst_campus is not None

        # ════════════════════════════════════════════════════════════════════
        # CASE A — both are campus / main-building locations
        #          → pure AR, no indoor path needed
        # ════════════════════════════════════════════════════════════════════
        if is_src_campus and is_dst_campus:
            src_key = src_campus
            dst_key = dst_campus
            src_gps = _CAMPUS_GPS.get(src_key)
            dst_gps = _CAMPUS_GPS.get(dst_key)

            if not src_gps:
                return False, (f"No GPS for '{src_display}'. "
                               "Check College_Db.xlsx GPS column.")
            if not dst_gps:
                return False, (f"No GPS for '{dst_display}'. "
                               "Check College_Db.xlsx GPS column.")

            self._src_lat, self._src_lon  = src_gps
            self._dst_lat, self._dst_lon  = dst_gps
            self._src_campus_key          = src_key
            self._dst_campus_key          = dst_key
            self.src_gps_display          = f"{src_gps[0]:.6f}, {src_gps[1]:.6f}"
            self.dst_gps_display          = f"{dst_gps[0]:.6f}, {dst_gps[1]:.6f}"
            self.static_distance_m        = _haversine(src_gps[0], src_gps[1],
                                                       dst_gps[0], dst_gps[1])
            initial_bearing               = _bearing(src_gps[0], src_gps[1],
                                                     dst_gps[0], dst_gps[1])
            compass_dir                   = _bearing_to_compass(initial_bearing)
            self.live_distance_m          = self.static_distance_m
            self.live_bearing_deg         = initial_bearing
            self.is_inter_building        = True
            self.is_navigating            = True
            self.ar_overlay_active        = True

            def _clean(v):
                return v.split("|", 1)[-1].strip() if "|" in v else v.strip()

            src_name = _clean(src_display).title()
            dst_name = _clean(dst_display).title()
            dist_str = (f"{self.static_distance_m/1000:.2f} km"
                        if self.static_distance_m >= 1000
                        else f"{self.static_distance_m:.0f} m")

            self._instructions = [
                f"Head {compass_dir} from {src_name} towards {dst_name}  (~{dist_str})",
                f"Keep walking towards {dst_name}. Stay on the path.",
                f"You have arrived at {dst_name}!",
            ]
            self.current_instruction_idx = 0
            print(f"[AR-A] {src_name} → {dst_name}  "
                  f"dist={self.static_distance_m:.1f} m  brg={initial_bearing:.1f}°")
            return True, "OK"

        # ════════════════════════════════════════════════════════════════════
        # CASE B + C — at least one side is an indoor room
        #   → resolve nodes, find graph path, compute indoor steps
        # ════════════════════════════════════════════════════════════════════
        self.is_inter_building = False

        src_node = (_extract_room_key(src_display)
                    if not is_src_campus else (src_campus or ""))
        dst_node = (_extract_room_key(dst_display)
                    if not is_dst_campus else (dst_campus or ""))

        missing = []
        if not src_node or src_node not in _GRAPH:
            missing.append(f"source '{src_display}'")
        if not dst_node or dst_node not in _GRAPH:
            missing.append(f"destination '{dst_display}'")
        if missing:
            return False, f"Could not find: {', '.join(missing)}"

        if src_node == dst_node:
            self._instructions           = ["You are already at your destination."]
            self.is_navigating           = True
            self.current_instruction_idx = 0
            return True, "OK"

        if not nx.has_path(_GRAPH, src_node, dst_node):
            return False, "No path found between selected rooms."

        try:
            self._path = nx.shortest_path(_GRAPH, src_node, dst_node,
                                          weight="weight")
        except nx.NetworkXNoPath:
            return False, "No path found between selected rooms."

        self.is_cross_block = any(n.startswith("CAMPUS_") for n in self._path)
        self._src_node               = src_node
        self._dst_node               = dst_node
        self._instructions           = _generate_instructions(self._path)
        self.current_instruction_idx = 0
        self.is_navigating           = True

        # ── GPS distance ─────────────────────────────────────────────────────
        # Determine the campus GPS key for each side.
        #   • campus locations  → their own key (already resolved above)
        #   • indoor rooms      → block-level pin (E_→CAMPUS_EAST, etc.)
        src_gps_key = src_campus if is_src_campus else _node_campus_key(src_node)
        dst_gps_key = dst_campus if is_dst_campus else _node_campus_key(dst_node)

        src_gps = _CAMPUS_GPS.get(src_gps_key) if src_gps_key else None
        dst_gps = _CAMPUS_GPS.get(dst_gps_key) if dst_gps_key else None

        # ── Determine if AR should be shown (cross-block rule) ───────────────
        # AR = ON  when the two GPS pins differ (different buildings/blocks).
        # AR = OFF when both pins are the same (same-block rooms → steps only).
        same_block = (src_gps_key == dst_gps_key)

        if src_gps and dst_gps:
            self._src_lat, self._src_lon  = src_gps
            self._dst_lat, self._dst_lon  = dst_gps
            self._src_campus_key          = src_gps_key
            self._dst_campus_key          = dst_gps_key
            self.src_gps_display          = f"{src_gps[0]:.6f}, {src_gps[1]:.6f}"
            self.dst_gps_display          = f"{dst_gps[0]:.6f}, {dst_gps[1]:.6f}"

            if same_block:
                # Same block: GPS distance = 0, use graph path weight instead
                graph_dist = sum(
                    _GRAPH.edges[self._path[k], self._path[k+1]].get("weight", 0)
                    for k in range(len(self._path) - 1)
                )
                self.static_distance_m = graph_dist
                self.live_bearing_deg  = None   # no meaningful bearing within same building
                self.ar_overlay_active = False  # NO AR for same-block navigation
                print(f"[Nav-C] same-block  path dist={graph_dist:.1f} m  NO AR")
            else:
                # Different blocks: show AR compass + distance
                self.static_distance_m = _haversine(src_gps[0], src_gps[1],
                                                    dst_gps[0], dst_gps[1])
                self.live_bearing_deg  = _bearing(src_gps[0], src_gps[1],
                                                  dst_gps[0], dst_gps[1])
                self.ar_overlay_active = True   # AR ON for cross-block navigation
                print(f"[Nav-B] cross-block  GPS dist={self.static_distance_m:.1f} m  "
                      f"brg={self.live_bearing_deg:.1f}°  AR=ON")

            self.live_distance_m = self.static_distance_m

        print(f"[Nav] {src_node} → {dst_node}  steps={len(self._instructions)}  "
              f"cross_block={self.is_cross_block}  same_gps={same_block}")
        return True, "OK"


    def get_all_instructions_text(self) -> List[str]:
        return list(self._instructions)

    def get_navigation_summary(self) -> dict:
        try:
            return self._get_navigation_summary_impl()
        except Exception as exc:
            print(f"[Nav] get_navigation_summary error: {exc}")
            return {"active": False}

    def _get_navigation_summary_impl(self) -> dict:
        if not self.is_navigating or not self._instructions:
            return {"active": False}
        idx   = self.current_instruction_idx
        total = len(self._instructions)
        text  = self._instructions[idx] if idx < total else ""

        current_floor = ""
        if self._path and idx < len(self._path):
            current_floor = _node_info(self._path[idx]).get("floor", "")

        d = {
            "active":        True,
            "current_step":  idx + 1,
            "total_steps":   total,
            "text":          text,
            "current_floor": current_floor,
            "is_inter_building": self.is_inter_building,
            "is_cross_block":    self.is_cross_block,
        }
        # Always expose GPS/distance fields — the screen uses static_distance_m
        # to display distance even in pure indoor (same-block) mode.
        # ar_overlay_active tells the screen whether to show the AR compass panel.
        compass = (_bearing_to_compass(self.live_bearing_deg)
                   if self.live_bearing_deg is not None else "")
        d["live_distance_m"]    = self.live_distance_m
        d["live_bearing_deg"]   = self.live_bearing_deg
        d["compass_direction"]  = compass
        d["gps_fixed"]          = self.gps_lat is not None
        d["src_gps"]            = self.src_gps_display
        d["dst_gps"]            = self.dst_gps_display
        d["static_distance_m"]  = self.static_distance_m
        d["src_campus_key"]     = self._src_campus_key
        d["dst_campus_key"]     = self._dst_campus_key
        d["ar_overlay_active"]  = self.ar_overlay_active
        return d

    def get_current_floor(self) -> str:
        return self.get_navigation_summary().get("current_floor", "")

    # ── AR frame overlay (unchanged) ─────────────────────────────────────────

    def render_ar_frame(self, frame):
        """
        Draw AR overlay on the camera frame.

        Layout
        ------
        TOP BAR   – current instruction text
        CENTER    – directional arrow
                    • GREEN  = live GPS lock
                    • AMBER  = static bearing from Excel coords (GPS locating)
                    • GREY   = no bearing data yet
        DIST STRIP– live distance (or Excel static estimate while locating)
        BOT BAR   – same instruction wrapped for readability

        GPS coordinates, bearing, and user position are shown in the
        KivyMD gps_overlay widget (top-right corner) — not re-drawn here.
        """
        # FIX: Use module-level cv2 import — do NOT re-import every frame
        try:
            if not CV2_AVAILABLE:
                return frame
            import cv2  # noqa: already imported at module level when available

            if not self.is_navigating or (not self.is_inter_building and not self.ar_overlay_active):
                return frame

            h, w   = frame.shape[:2]
            cx, cy = w // 2, h // 2
            gps_ok = (self.gps_lat is not None and
                      self.live_bearing_deg is not None and
                      self.live_distance_m is not None)
            al = min(w, h) // 3   # arrow length

            # ── Directional arrow ─────────────────────────────────────────────
            bearing = self.live_bearing_deg   # set to static value at start
            if bearing is not None:
                tail, head = _arrow_offset(bearing, self.compass_heading, cx, cy, al)
                arrow_col  = (0, 230, 80) if gps_ok else (255, 200, 60)
            else:
                tail = (cx, cy + al // 2)
                head = (cx, cy - al // 2)
                arrow_col = (180, 180, 180)

            tip = 0.30
            cv2.arrowedLine(frame, tail, head, (0, 0, 0),
                            max(18, al // 8), tipLength=tip)
            cv2.arrowedLine(frame, tail, head, arrow_col,
                            max(12, al // 12), tipLength=tip)

            # ── TOP BAR ───────────────────────────────────────────────────────
            top_h = max(55, h // 9)
            ov = frame.copy()
            cv2.rectangle(ov, (0, 0), (w, top_h), (8, 8, 24), -1)
            cv2.addWeighted(ov, 0.82, frame, 0.18, 0, frame)

            idx  = self.current_instruction_idx
            step = self._instructions[idx] if idx < len(self._instructions) else ""
            cv2.putText(frame, step[:48],
                        (12, int(top_h * 0.62)),
                        cv2.FONT_HERSHEY_DUPLEX, 0.50,
                        (220, 220, 255), 1, cv2.LINE_AA)

            # ── DISTANCE + COMPASS STRIP ──────────────────────────────────────
            # Prefer live GPS distance; fall back to static Excel distance
            dist_m = self.live_distance_m if gps_ok else self.static_distance_m
            if dist_m is not None:
                dist_str  = (f"{dist_m/1000:.2f} km" if dist_m >= 1000
                             else f"{dist_m:.0f} m")
                dist_col  = (60, 240, 110) if gps_ok else (255, 200, 60)
                est_label = "" if gps_ok else "  est."
                full_dist = dist_str + est_label
            else:
                full_dist = "GPS locating..."
                dist_col  = (60, 210, 210)

            s1 = top_h + 4
            s2 = s1 + int(h * 0.09) + 10
            ov2 = frame.copy()
            cv2.rectangle(ov2, (0, s1), (w, s2), (0, 0, 0), -1)
            cv2.addWeighted(ov2, 0.45, frame, 0.55, 0, frame)

            # Scale distance font to frame height; ~30% smaller than before
            _dist_scale = max(0.65, min(1.3, h / 240))
            _dist_thick = max(1, int(_dist_scale * 1.2))

            (tw, _), _ = cv2.getTextSize(full_dist, cv2.FONT_HERSHEY_DUPLEX, _dist_scale, _dist_thick)
            cv2.putText(frame, full_dist,
                        (max(10, (w - tw) // 2), s2 - 6),
                        cv2.FONT_HERSHEY_DUPLEX, _dist_scale, dist_col, _dist_thick, cv2.LINE_AA)

            # Compass label — shown below the distance text, centred, smaller
            # (previously placed at w-140 which overlapped the distance text)
            if bearing is not None:
                compass_lbl = _bearing_to_compass(bearing)
                bear_txt    = f"{compass_lbl}  {bearing:.0f}\u00b0"
                (btw, _), _ = cv2.getTextSize(
                    bear_txt, cv2.FONT_HERSHEY_DUPLEX, 0.42, 1)
                cv2.putText(frame, bear_txt,
                            (max(10, (w - btw) // 2), s2 + 14),
                            cv2.FONT_HERSHEY_DUPLEX, 0.42,
                            (180, 180, 240), 1, cv2.LINE_AA)

            # ── BOTTOM BAR ────────────────────────────────────────────────────
            # REMOVED: bottom bar was re-rendering the same `step` text that
            # already appears in the TOP BAR, causing duplicate text on screen.
            # The Kivy HUD bar (instr_lbl) also shows the step — three copies
            # of the same string is what caused the visible overlap. Removed.

        except Exception as exc:
            print(f"[AR] render_ar_frame error: {exc}")
        return frame


# ─────────────────────────────────────────────────────────────────────────────
#  SELF-TEST
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import os

    # ── Load real GPS from Excel first ────────────────────────────────────────
    EXCEL = os.path.join(os.path.dirname(__file__), "College_Db.xlsx")
    if not os.path.isfile(EXCEL):
        EXCEL = "/mnt/user-data/uploads/College_Db.xlsx"   # dev fallback
    reloaded = reload_campus_gps(EXCEL)
    print(f"GPS loaded from Excel: {reloaded}")
    print()

    # ── TEST 1: AR – North Block main building → East Block main building ─────
    print("=" * 68)
    print("TEST 1 (AR): North Block main bldg → East Block main bldg")
    print("=" * 68)
    c1 = NavigationController()
    ok1, msg1 = c1.start_navigation(
        "NORTH BLOCK|North Block Main building",
        "EAST BLOCK|East block main Building",
    )
    print(f"  ok={ok1}  msg={msg1!r}")
    s1 = c1.get_navigation_summary()
    print(f"  SRC GPS  : {s1['src_gps']}")
    print(f"  DST GPS  : {s1['dst_gps']}")
    print(f"  Distance : {s1['static_distance_m']:.1f} m")
    print(f"  Bearing  : {s1['live_bearing_deg']:.1f}°  ({s1['compass_direction']})")
    for step in c1.get_all_instructions_text():
        print(f"  → {step}")

    # ── TEST 2: AR – Gate-1 → Sports Ground ──────────────────────────────────
    print()
    print("=" * 68)
    print("TEST 2 (AR): Gate-1 → Sports Ground")
    print("=" * 68)
    c2 = NavigationController()
    ok2, msg2 = c2.start_navigation("GATE-1", "SPORTS GROUND")
    print(f"  ok={ok2}  msg={msg2!r}")
    s2 = c2.get_navigation_summary()
    print(f"  SRC GPS  : {s2['src_gps']}")
    print(f"  DST GPS  : {s2['dst_gps']}")
    print(f"  Distance : {s2['static_distance_m']:.1f} m")
    print(f"  Bearing  : {s2['live_bearing_deg']:.1f}°  ({s2['compass_direction']})")
    for step in c2.get_all_instructions_text():
        print(f"  → {step}")

    # ── TEST 3: STEPS – North 1F Computer Lab-2 → South 3F IT HOD ────────────
    print()
    print("=" * 68)
    print("TEST 3 (STEPS): N-Block Computer Lab-2 (107) → S-Block IT HOD (302)")
    print("=" * 68)
    c3 = NavigationController()
    ok3, msg3 = c3.start_navigation(
        "NORTH BLOCK|Computer Lab-2 - 107",
        "SOUTH BLOCK|IT HOD - 302",
    )
    print(f"  ok={ok3}  msg={msg3!r}")
    for idx, s in enumerate(c3.get_all_instructions_text(), 1):
        print(f"  Step {idx:>2}: {s}")

    # ── TEST 4: AR live GPS simulation ───────────────────────────────────────
    print()
    print("=" * 68)
    print("TEST 4 (AR GPS simulation): West Block → Bench Park")
    print("=" * 68)
    c4 = NavigationController()
    c4.start_navigation("WEST BLOCK", "BENCH PARK")
    # Simulate phone standing at West Block GPS pin
    west_lat, west_lon = _CAMPUS_GPS["CAMPUS_WEST"]
    c4.on_gps_update(west_lat, west_lon)
    c4.on_compass_update(0.0)   # facing North
    s4 = c4.get_navigation_summary()
    print(f"  GPS fixed        : {s4['gps_fixed']}")
    print(f"  Live distance    : {s4['live_distance_m']:.1f} m")
    print(f"  Live bearing     : {s4['live_bearing_deg']:.1f}°  ({s4['compass_direction']})")
    print(f"  SRC GPS          : {s4['src_gps']}")
    print(f"  DST GPS          : {s4['dst_gps']}")