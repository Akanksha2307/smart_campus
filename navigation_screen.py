"""
navigation_screen.py  -  SmartCampus  (Campus Navigation)
==========================================================

ANDROID CAMERA FIX SUMMARY (v2 - Android 15 Compatible)
---------------------------------------------------------
ROOT CAUSE ANALYSIS:
1. OpenCV (cv2) VideoCapture does NOT work reliably on Android via Buildozer.
   - Android restricts direct /dev/videoN access from Python processes.
   - cv2.VideoCapture always returns isOpened()=True but frames are black/None.
   - This is a known limitation of opencv-python on Android ARM targets.

2. Android 6+ requires RUNTIME permission grants (not just manifest).
   - Android 15 (API 35) enforces stricter permission scoping.
   - Permissions must be requested AFTER the Activity window is visible.
   - Silent camera failure = permission not granted at the moment of capture.

3. The correct Android camera approach for Kivy/Buildozer is:
   - Use Kivy's built-in Camera widget (uses Android Camera2 API via Java)
   - OR use Plyer's camera module (safe wrapper around Android Camera API)
   - NOT cv2.VideoCapture (unreliable on Android, works only on desktop/Linux)

FIXES APPLIED:
1. Primary camera = Kivy Camera widget (works natively on Android).
2. cv2.VideoCapture = FALLBACK only (desktop/Linux testing only).
3. Permissions: requested via callback at on_enter() (Activity is visible).
4. Android 15: added READ_MEDIA_IMAGES + POST_NOTIFICATIONS permissions.
5. Camera warm-up delay of 0.5s before first frame read.
6. Multiple camera index fallback (0, 1, -1) with 0.2s delay each.
7. AR overlay renders even without camera (Kivy compass panel = primary UI).
8. Added _use_kivy_camera flag to switch provider at runtime.
"""

import os
import time

# ── cv2 is optional: available on desktop, gracefully absent on Android ──────
try:
    import cv2
    CV2_AVAILABLE = True
except ImportError:
    cv2 = None
    CV2_AVAILABLE = False
    print("[Nav] cv2 not available – camera features disabled (Android mode)")

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    openpyxl = None
    _OPENPYXL_OK = False
    print("[Nav] openpyxl not available – room lists will use node fallback")

from kivy.clock            import Clock
from kivy.core.window      import Window
from kivy.graphics         import Color, Rectangle, RoundedRectangle
from kivy.graphics.texture import Texture
from kivy.metrics          import dp, sp
from kivy.uix.boxlayout    import BoxLayout
from kivy.uix.floatlayout  import FloatLayout
from kivy.uix.relativelayout import RelativeLayout
from kivy.uix.image        import Image
from kivy.uix.modalview    import ModalView
from kivy.uix.scrollview   import ScrollView

# ── Kivy Camera widget (primary on Android) ───────────────────────────────────
try:
    from kivy.uix.camera import Camera as KivyCamera
    KIVY_CAMERA_AVAILABLE = True
except ImportError:
    KivyCamera = None
    KIVY_CAMERA_AVAILABLE = False
    print("[Nav] Kivy Camera widget not available")

# ── Lock to PORTRAIT orientation ─────────────────────────────────────────────
try:
    from jnius import autoclass                        # type: ignore
    ActivityInfo   = autoclass("android.content.pm.ActivityInfo")
    PythonActivity = autoclass("org.kivy.android.PythonActivity")
    PythonActivity.mActivity.setRequestedOrientation(
        ActivityInfo.SCREEN_ORIENTATION_PORTRAIT
    )
    print("[Nav] Android orientation locked to PORTRAIT")
except Exception:
    pass

Window.softinput_mode = "below_target"

from kivymd.uix.screen    import MDScreen
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.card      import MDCard
from kivymd.uix.button    import MDRaisedButton
from kivymd.uix.label     import MDLabel
from kivymd.uix.toolbar   import MDTopAppBar

from navigation_controller import (NavigationController, reload_campus_gps,
                                   get_campus_gps, _CAMPUS_NAME_MAP)

# ============================================================================
#  Campus building/location list
# ============================================================================

_BLOCK_SHEETS = {
    "East":  "EAST BLOCK",
    "North": "NORTH BLOCK",
    "South": "SOUTH BLOCK",
    "West":  "WEST BLOCK",
}

CAMPUS_BUILDINGS = [
    "EAST BLOCK",
    "SOUTH BLOCK",
    "NORTH BLOCK",
    "WEST BLOCK",
    "BENCH PARK",
    "CANTEEN & STATIONARY",
    "EWS WORKSHOP",
    "GATE-1",
    "GATE-2 ENTRY",
    "MOBIKES PARKING",
    "OPEN AIR AUDITORIUM",
    "SPORTS GROUND",
]
CAMPUS_BUILDINGS_SET = set(CAMPUS_BUILDINGS)

_MAIN_BLDG_ROOMS = {
    "East block main Building",
    "North Block Main building",
    "South Block main Building",
    "West block main building",
}

_MAIN_BLDG_KEYS = (
    {f"EAST BLOCK|{r}"  for r in _MAIN_BLDG_ROOMS} |
    {f"NORTH BLOCK|{r}" for r in _MAIN_BLDG_ROOMS} |
    {f"SOUTH BLOCK|{r}" for r in _MAIN_BLDG_ROOMS} |
    {f"WEST BLOCK|{r}"  for r in _MAIN_BLDG_ROOMS}
)

AR_LOCATIONS_SET = CAMPUS_BUILDINGS_SET | _MAIN_BLDG_KEYS

_SUB_LABEL = {
    "EAST BLOCK":           "East Block Building",
    "NORTH BLOCK":          "North Block Building",
    "SOUTH BLOCK":          "South Block Building",
    "WEST BLOCK":           "West Block Building",
    "BENCH PARK":           "Bench Park Area",
    "CANTEEN & STATIONARY": "Canteen & Stationary Shop",
    "EWS WORKSHOP":         "EWS Workshop",
    "GATE-1":               "Gate 1 – Main Entrance",
    "GATE-2 ENTRY":         "Gate 2 – Entry",
    "MOBIKES PARKING":      "Motorbikes Parking Area",
    "OPEN AIR AUDITORIUM":  "Open Air Auditorium",
    "SPORTS GROUND":        "Sports Ground",
}

_BLD_COLOR = {
    "EAST BLOCK":           (0.05, 0.50, 0.38),
    "SOUTH BLOCK":          (0.05, 0.42, 0.78),
    "NORTH BLOCK":          (0.52, 0.18, 0.68),
    "WEST BLOCK":           (0.70, 0.34, 0.05),
    "BENCH PARK":           (0.18, 0.52, 0.10),
    "CANTEEN & STATIONARY": (0.72, 0.52, 0.08),
    "EWS WORKSHOP":         (0.55, 0.28, 0.08),
    "GATE-1":               (0.38, 0.38, 0.44),
    "GATE-2 ENTRY":         (0.38, 0.38, 0.44),
    "MOBIKES PARKING":      (0.28, 0.28, 0.32),
    "OPEN AIR AUDITORIUM":  (0.05, 0.52, 0.60),
    "SPORTS GROUND":        (0.78, 0.28, 0.08),
}
_DEF_COLOR = (0.15, 0.15, 0.30)


def _bcolor(name: str):
    return _BLD_COLOR.get(name.upper(), _DEF_COLOR)


# ============================================================================
#  Excel loader
# ============================================================================

def _fmt_room(name: str, number) -> str:
    num = str(number).strip() if number is not None else ""
    if num and num != "-":
        return f"{name} - {num}"
    return name


def _find_excel() -> str:
    try:
        here = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        here = os.getcwd()

    cwd = os.getcwd()

    fnames = [
        "College_Db.xlsx",
        "college_db.xlsx",
        "College_DB.xlsx",
        "Campus_Room_Directory.xlsx",
        "Campus_Room_Directory.excel",
        "campus_room_directory.xlsx",
        "campus_room_directory.excel",
    ]

    search_dirs = []
    for base in [here, cwd]:
        search_dirs += [
            base,
            os.path.join(base, "data"),
            os.path.join(base, "data", "listofRooms"),
            os.path.join(base, "assets"),
            os.path.join(base, ".."),
            os.path.join(base, "..", "data"),
            os.path.join(base, "..", "data", "listofRooms"),
            os.path.join(base, "..", "assets"),
            os.path.join(base, "..", ".."),
            os.path.join(base, "..", "..", "data"),
            os.path.join(base, "..", "..", "data", "listofRooms"),
        ]

    # Android-specific paths
    try:
        from android.storage import app_storage_path  # type: ignore
        android_base = app_storage_path()
        search_dirs += [
            android_base,
            os.path.join(android_base, "files"),
            os.path.join(android_base, "assets"),
        ]
    except Exception:
        pass

    seen, unique_dirs = set(), []
    for d in search_dirs:
        d = os.path.normpath(d)
        if d not in seen:
            seen.add(d)
            unique_dirs.append(d)

    for d in unique_dirs:
        for fn in fnames:
            p = os.path.join(d, fn)
            if os.path.isfile(p):
                print(f"[Nav] Excel found: {p}")
                return p

    print(f"[Nav] Excel NOT found. cwd={cwd}")
    return ""


def _load_excel(path: str) -> dict:
    if not _OPENPYXL_OK or openpyxl is None:
        print("[Nav] openpyxl unavailable - returning empty room list")
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"[Nav] Excel open error: {exc}")
        return {}

    result: dict = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        block_key = _BLOCK_SHEETS.get(sheet_name)
        if block_key:
            rooms = []
            for row in rows[2:]:
                if not any(row):
                    continue
                name   = row[1] if len(row) > 1 else None
                number = row[2] if len(row) > 2 else None

                if name is None:
                    continue
                name_str = str(name).strip()
                if not name_str:
                    continue

                rooms.append(_fmt_room(name_str, number))

            result[block_key] = rooms
            print(f"[Nav] {block_key}: {len(rooms)} rooms loaded")
            continue

        if sheet_name == "Main_buildingNames":
            for row in rows[1:]:
                if not any(row):
                    continue
                bname = row[1] if len(row) > 1 else None
                if bname is None:
                    continue
                bname_str = str(bname).strip()
                if not bname_str:
                    continue
                key = bname_str.upper()
                if key not in _BLOCK_SHEETS.values():
                    result.setdefault(key, [])
            continue

    wb.close()

    for bld in CAMPUS_BUILDINGS:
        result.setdefault(bld, [])

    return result


def _load_from_nodes() -> dict:
    try:
        from navigation_controller import _NODES

        result = {
            "SOUTH BLOCK": [],
            "NORTH BLOCK": [],
            "EAST BLOCK":  [],
            "WEST BLOCK":  [],
        }
        block_map = {"S": "SOUTH BLOCK", "N": "NORTH BLOCK",
                     "E": "EAST BLOCK",  "W": "WEST BLOCK"}

        for rid, attrs in _NODES.items():
            blk_char = attrs.get("block", "")
            if not blk_char:
                blk_char = rid[0] if rid and rid[0] in "SNEW" else ""
            blk_key = block_map.get(blk_char)
            if blk_key is None:
                continue
            name    = attrs.get("name", rid)
            rnum    = attrs.get("room_num", "")
            label   = f"{name} - {rnum}" if rnum and rnum not in ("-", "") else name
            result[blk_key].append(label)

        for k in result:
            result[k].sort()
        return result
    except Exception:
        return {}


def _build_room_data() -> dict:
    path  = _find_excel()
    rooms = _load_excel(path) if path else {}
    if not any(rooms.values()):
        rooms = _load_from_nodes()
    for bld in CAMPUS_BUILDINGS:
        rooms.setdefault(bld, [])
    return rooms


# ============================================================================
#  Android runtime permissions
# ============================================================================
import traceback

_CAMERA_GRANTED   = False
_LOCATION_GRANTED = False
_IS_ANDROID       = False

try:
    from android.permissions import (          # type: ignore
        request_permissions, check_permission, Permission
    )
    _IS_ANDROID = True

    def _on_permissions_result(permissions, grants):
        global _CAMERA_GRANTED, _LOCATION_GRANTED
        for perm, granted in zip(permissions, grants):
            if perm == Permission.CAMERA:
                _CAMERA_GRANTED = granted
                print(f"[Nav] CAMERA permission: {'GRANTED' if granted else 'DENIED'}")
            if perm in (Permission.ACCESS_FINE_LOCATION,
                        Permission.ACCESS_COARSE_LOCATION):
                _LOCATION_GRANTED = _LOCATION_GRANTED or granted
        print(f"[Nav] Permissions → Camera={_CAMERA_GRANTED}, Location={_LOCATION_GRANTED}")

    # FIX: Only check current status at module load — do NOT call
    # request_permissions() here.  Android 6+ (and especially Android 15)
    # silently drops permission dialogs that are raised before the Activity
    # window is fully visible.  The actual request is made in on_enter()
    # where the screen is guaranteed to be in the foreground.
    try:
        _CAMERA_GRANTED   = check_permission(Permission.CAMERA)
        _LOCATION_GRANTED = check_permission(Permission.ACCESS_FINE_LOCATION)
        print(f"[Nav] Initial check — Camera={_CAMERA_GRANTED}, Location={_LOCATION_GRANTED}")
    except Exception:
        pass

except Exception:
    # Desktop / non-Android — all permissions implicitly available
    _CAMERA_GRANTED   = True
    _LOCATION_GRANTED = True

try:
    from database import Database
except ImportError:
    class Database:  # type: ignore[no-redef]
        def log_error(self, *args, **kwargs):
            pass

_NAV_DB = None
def _get_db():
    global _NAV_DB
    if _NAV_DB is None:
        try:
            _NAV_DB = Database()
        except Exception:
            pass
    return _NAV_DB


# ============================================================================
#  NavPopup
# ============================================================================

class NavPopup(ModalView):

    def __init__(self, caller, room_data: dict, on_pick, **kwargs):
        super().__init__(
            size_hint=(1, 1),
            background_color=(0, 0, 0, 0),
            background="",
            overlay_color=(0, 0, 0, 0.35),
            auto_dismiss=True,
            **kwargs
        )
        self._caller    = caller
        self._on_pick   = on_pick
        self._room_data = room_data
        self._build()
        Window.bind(on_resize=self._on_window_resize)

    def on_dismiss(self):
        Window.unbind(on_resize=self._on_window_resize)

    def _on_window_resize(self, win, w, h):
        self.clear_widgets()
        self._build()

    def _calc_popup_geometry(self):
        try:
            cx, cy = self._caller.to_window(*self._caller.pos)
        except Exception:
            cx = getattr(self._caller, "x", 0)
            cy = getattr(self._caller, "y", 0)

        cw = getattr(self._caller, "width",  Window.width)
        ch = getattr(self._caller, "height", dp(44))

        popup_w = min(max(cw, dp(260)), Window.width - dp(12))
        popup_h = min(dp(500), Window.height * 0.78)

        popup_x = cx
        popup_y = cy - popup_h

        if popup_y < dp(8):
            popup_y = cy + ch
            if popup_y + popup_h > Window.height - dp(8):
                popup_y = max(dp(8), Window.height - popup_h - dp(8))

        if popup_x + popup_w > Window.width - dp(8):
            popup_x = Window.width - popup_w - dp(8)
        if popup_x < dp(8):
            popup_x = dp(8)

        return popup_x, popup_y, popup_w, popup_h

    def _build(self):
        popup_x, popup_y, popup_w, popup_h = self._calc_popup_geometry()

        outer = FloatLayout(size_hint=(1, 1))
        self.add_widget(outer)

        card = MDCard(
            orientation="vertical",
            size_hint=(None, None),
            size=(popup_w, popup_h),
            pos=(popup_x, popup_y),
            elevation=16,
            md_bg_color=(1, 1, 1, 1),
            radius=[dp(12)],
        )
        outer.add_widget(card)

        from kivy.uix.textinput import TextInput
        search_bar = TextInput(
            hint_text="Search location or room…",
            multiline=False,
            size_hint=(1, None),
            height=dp(40),
            font_size=sp(12),
            padding=[dp(10), dp(10)],
            foreground_color=(0.1, 0.1, 0.2, 1),
            background_color=(0.94, 0.96, 1.0, 1),
            cursor_color=(0.1, 0.4, 0.8, 1),
        )
        card.add_widget(search_bar)

        sv = ScrollView(
            size_hint=(1, 1),
            bar_width=dp(5),
            bar_color=[0.2, 0.5, 0.9, 0.9],
            bar_inactive_color=[0.2, 0.5, 0.9, 0.3],
            scroll_type=["bars", "content"],
            do_scroll_x=False,
        )
        card.add_widget(sv)

        self._content = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            spacing=0,
        )
        self._content.bind(minimum_height=self._content.setter("height"))
        sv.add_widget(self._content)

        self._all_rows = []
        for bld in CAMPUS_BUILDINGS:
            r, g, b = _bcolor(bld)
            rooms   = self._room_data.get(bld, [])
            hdr = self._bld_row(bld, r, g, b)
            self._content.add_widget(hdr)
            self._all_rows.append((bld.lower(), hdr))

            if rooms:
                for rname in rooms:
                    rw = self._room_row(rname, bld, r, g, b)
                    self._content.add_widget(rw)
                    self._all_rows.append((f"{bld} {rname}".lower(), rw))
            else:
                sub = self._sub_row(_SUB_LABEL.get(bld, bld.title()), r, g, b)
                self._content.add_widget(sub)
                self._all_rows.append((bld.lower(), sub))

        search_bar.bind(text=self._on_search)

    def _on_search(self, instance, text):
        query = text.strip().lower()
        for key, widget in self._all_rows:
            widget.opacity = 1 if (not query or query in key) else 0
            if hasattr(widget, "_base_height"):
                widget.height = 0 if widget.opacity == 0 else widget._base_height

    def _bld_row(self, name: str, r, g, b):
        ROW_H = dp(48)
        from kivy.uix.behaviors import ButtonBehavior

        class _TapBox(ButtonBehavior, BoxLayout):
            pass

        row = _TapBox(
            orientation="horizontal",
            size_hint=(1, None), height=ROW_H,
        )
        row._base_height = ROW_H

        strip = BoxLayout(size_hint=(None, 1), width=dp(6))
        with strip.canvas.before:
            Color(r, g, b, 1)
            strip._rect = Rectangle(pos=strip.pos, size=strip.size)
        strip.bind(
            pos =lambda w, v: setattr(w._rect, "pos",  v),
            size=lambda w, v: setattr(w._rect, "size", v),
        )
        row.add_widget(strip)

        with row.canvas.before:
            Color(0.92 + r * 0.04, 0.92 + g * 0.04, 0.92 + b * 0.04, 1)
            row._bg = Rectangle(pos=row.pos, size=row.size)
        row.bind(
            pos =lambda w, v: setattr(w._bg, "pos",  v),
            size=lambda w, v: setattr(w._bg, "size", v),
        )

        display = (name.title()
                   .replace("Ews", "EWS")
                   .replace("Gate-1", "Gate-1")
                   .replace("Gate-2 Entry", "Gate-2 Entry"))
        lbl = MDLabel(
            text=display,
            font_size=sp(12),
            bold=True,
            halign="left",
            valign="middle",
            theme_text_color="Primary",
            padding=[dp(10), 0],
        )
        lbl.bind(size=lbl.setter("text_size"))
        row.add_widget(lbl)

        def _on_release(inst):
            self._on_pick(name)
            self.dismiss()
        row.bind(on_release=_on_release)
        return row

    def _room_row(self, name: str, bld: str, r, g, b):
        ROW_H = dp(44)
        from kivy.uix.behaviors import ButtonBehavior

        class _TapBox(ButtonBehavior, BoxLayout):
            pass

        row = _TapBox(
            orientation="horizontal",
            size_hint=(1, None), height=ROW_H,
            padding=[dp(18), 0, dp(6), 0],
        )
        row._base_height = ROW_H

        with row.canvas.before:
            Color(1, 1, 1, 1)
            row._bg = Rectangle(pos=row.pos, size=row.size)
        row.bind(
            pos =lambda w, v: setattr(w._bg, "pos",  v),
            size=lambda w, v: setattr(w._bg, "size", v),
        )

        lbl = MDLabel(
            text=name,
            font_size=sp(11),
            halign="left",
            valign="middle",
            theme_text_color="Custom",
            text_color=(r * 0.55, g * 0.55, b * 0.65, 0.92),
        )
        lbl.bind(size=lbl.setter("text_size"))
        row.add_widget(lbl)

        def _on_release(inst):
            self._on_pick(f"{bld}|{name}")
            self.dismiss()
        row.bind(on_release=_on_release)
        return row

    def _sub_row(self, name: str, r, g, b):
        ROW_H = dp(30)
        row = BoxLayout(
            orientation="horizontal",
            size_hint=(1, None), height=ROW_H,
            padding=[dp(16), 0],
        )
        row._base_height = ROW_H
        with row.canvas.before:
            Color(r, g, b, 0.06)
            row._bg = Rectangle(pos=row.pos, size=row.size)
        row.bind(
            pos =lambda w, v: setattr(w._bg, "pos",  v),
            size=lambda w, v: setattr(w._bg, "size", v),
        )
        lbl = MDLabel(
            text=name,
            font_size=sp(10),
            halign="left",
            valign="middle",
            theme_text_color="Custom",
            text_color=(r, g, b, 0.72),
        )
        lbl.bind(size=lbl.setter("text_size"))
        row.add_widget(lbl)
        return row


# ============================================================================
#  Dropdown trigger button
# ============================================================================

class NavDropBtn(MDBoxLayout):

    def __init__(self, placeholder: str,
                 accent=(0.05, 0.42, 0.78),
                 **kwargs):
        super().__init__(
            orientation="horizontal", size_hint=(1, None), height=dp(52),
            padding=[dp(12), dp(6)], spacing=dp(8), **kwargs
        )
        self._placeholder = placeholder
        self._value       = ""
        self._accent      = accent
        self._empty_rgba  = (0.93, 0.93, 0.97, 1)
        self._filled_rgba = (*accent, 0.12)

        with self.canvas.before:
            self._bg_col  = Color(*self._empty_rgba)
            self._bg_rect = RoundedRectangle(
                pos=self.pos, size=self.size, radius=[dp(10)])
            self._dot_col  = Color(0, 0, 0, 0)
            self._dot_rect = RoundedRectangle(
                size=(dp(4), dp(24)), radius=[dp(2)])
        self.bind(pos=self._upd, size=self._upd)

        self._lbl = MDLabel(
            text=placeholder, font_size=sp(13),
            theme_text_color="Hint", halign="left",
        )
        self._chev = MDLabel(
            text="▼", font_size=sp(10), bold=True,
            size_hint=(None, 1), width=dp(20),
            halign="right", theme_text_color="Hint",
        )
        self.add_widget(self._lbl)
        self.add_widget(self._chev)
        self.on_tap = lambda: None

    def _upd(self, *_):
        self._bg_rect.pos  = self.pos
        self._bg_rect.size = self.size
        self._dot_rect.pos = (self.x + dp(2),
                              self.y + (self.height - dp(24)) / 2)

    def set_value(self, val: str):
        self._value = val
        if val:
            display_raw = val.split("|", 1)[-1] if "|" in val else val
            display     = display_raw if len(display_raw) <= 34 else display_raw[:32] + "…"
            self._lbl.text              = display
            self._lbl.font_size         = sp(12)
            self._lbl.theme_text_color  = "Primary"
            self._bg_col.rgba           = self._filled_rgba
            self._dot_col.rgba          = (*self._accent, 1)
            self._chev.theme_text_color = "Primary"
        else:
            self._lbl.text              = self._placeholder
            self._lbl.font_size         = sp(13)
            self._lbl.theme_text_color  = "Hint"
            self._bg_col.rgba           = self._empty_rgba
            self._dot_col.rgba          = (0, 0, 0, 0)
            self._chev.theme_text_color = "Hint"

    @property
    def text(self):
        return self._value

    def on_touch_down(self, touch):
        if self.collide_point(*touch.pos):
            self._bg_col.rgba = (*self._accent[:3], 0.22)
            return True
        return super().on_touch_down(touch)

    def on_touch_up(self, touch):
        if self.collide_point(*touch.pos):
            if self._value:
                self._bg_col.rgba = self._filled_rgba
            else:
                self._bg_col.rgba = self._empty_rgba
            self.on_tap()
            return True
        return super().on_touch_up(touch)


# ============================================================================
#  Step card helper
# ============================================================================

def _make_step_card(idx: int, step: str):
    t = step.lower()
    if   "lift"     in t:                       tag, r, g, b = "LFT", 0.85, 0.52, 0.05
    elif "stair"    in t:                       tag, r, g, b = "STR", 0.52, 0.20, 0.70
    elif "arrived"  in t:                       tag, r, g, b = "END", 0.05, 0.60, 0.32
    elif "washroom" in t or " wc" in t:        tag, r, g, b = "WC",  0.05, 0.50, 0.60
    elif "exit"     in t or "you are at" in t: tag, r, g, b = "OUT", 0.18, 0.44, 0.82
    elif "walk"     in t or "corridor" in t:   tag, r, g, b = "GO",  0.15, 0.42, 0.82
    else:                                       tag, r, g, b = ">>",  0.15, 0.42, 0.82

    card = MDCard(
        orientation="horizontal",
        size_hint=(1, None), height=dp(60),
        padding=0, spacing=0,
        elevation=2, md_bg_color=(1, 1, 1, 1),
        radius=[dp(8)],
    )
    card.add_widget(MDBoxLayout(
        size_hint=(None, 1), width=dp(5),
        md_bg_color=(r, g, b, 1),
    ))
    inner = MDBoxLayout(
        orientation="horizontal", size_hint=(1, 1),
        padding=[dp(8), dp(4)], spacing=dp(6),
    )
    bubble = MDBoxLayout(
        size_hint=(None, None), size=(dp(26), dp(26)),
        md_bg_color=(r, g, b, 0.15), radius=[dp(13)],
    )
    bubble.add_widget(MDLabel(
        text=str(idx), font_size=sp(9), bold=True,
        halign="center", valign="middle",
        theme_text_color="Custom", text_color=(r, g, b, 1),
    ))
    bwrap = BoxLayout(size_hint=(None, 1), width=dp(30), orientation="vertical")
    bwrap.add_widget(BoxLayout())
    bwrap.add_widget(bubble)
    bwrap.add_widget(BoxLayout())

    tl   = MDLabel(text=tag, font_size=sp(8), bold=True,
                   size_hint=(None, 1), width=dp(26), halign="center",
                   theme_text_color="Custom", text_color=(r, g, b, 1))
    desc = MDLabel(text=step, font_size=sp(11),
                   halign="left", valign="middle", theme_text_color="Primary")
    desc.bind(size=desc.setter("text_size"))

    inner.add_widget(bwrap)
    inner.add_widget(tl)
    inner.add_widget(desc)
    card.add_widget(inner)
    return card


# ============================================================================
#  Main Screen
# ============================================================================

class NavigationScreen(MDScreen):

    def __init__(self, **kwargs):
        try:
            super().__init__(**kwargs)
            self.controller    = None
            self.cap           = None            # cv2 VideoCapture (desktop only)
            self._kivy_cam     = None            # Kivy Camera widget (Android)
            self._use_kivy_cam = _IS_ANDROID     # prefer Kivy cam on Android
            self._room_data: dict = {}
            self._active_popup = None
            self._cam_small    = False
            self._first_frame  = True

            root = BoxLayout(orientation="vertical", size_hint=(1, 1),
                             pos_hint={"top": 1})
            self.add_widget(root)

            root.add_widget(MDTopAppBar(
                title="Campus Navigation",
                anchor_title="center",
                left_action_items=[["arrow-left", lambda x: self._on_back()]],
                md_bg_color=(0.10, 0.10, 0.20, 1),
                elevation=2,
            ))

            sel = MDCard(
                orientation="vertical", padding=dp(12), spacing=dp(8),
                size_hint=(1, None), height=dp(220),
                radius=[0, 0, dp(18), dp(18)],
                elevation=4, md_bg_color=(1, 1, 1, 1),
            )
            sel.add_widget(MDLabel(
                text="NAVIGATE TO", bold=True, font_style="Caption",
                theme_text_color="Secondary",
                size_hint_y=None, height=dp(16),
            ))

            self.src_btn = NavDropBtn("Select Source",      accent=(0.05, 0.42, 0.78))
            self.dst_btn = NavDropBtn("Select Destination", accent=(0.52, 0.18, 0.68))
            sel.add_widget(self.src_btn)
            sel.add_widget(self.dst_btn)
            sel.add_widget(MDRaisedButton(
                text="START NAVIGATION",
                size_hint=(1, None), height=dp(44),
                md_bg_color=(0.10, 0.10, 0.20, 1),
                on_release=self.start_nav,
            ))
            root.add_widget(sel)

            self.err_lbl = MDLabel(
                text="", theme_text_color="Error",
                halign="center", size_hint_y=None,
                height=dp(22), font_size=sp(10.5),
            )
            root.add_widget(self.err_lbl)

            nav = BoxLayout(orientation="vertical", size_hint=(1, 1))
            self.hud = MDBoxLayout(
                size_hint=(1, None), height=dp(42),
                md_bg_color=(0.10, 0.10, 0.22, 1),
                opacity=0, padding=[dp(10), 0],
            )
            self.instr_lbl = MDLabel(
                text="", halign="center",
                theme_text_color="Custom", text_color=(1, 1, 1, 1),
                font_size=sp(10.5), bold=True,
            )
            self.hud.add_widget(self.instr_lbl)
            nav.add_widget(self.hud)
            self.content_box = BoxLayout(size_hint=(1, 1))
            nav.add_widget(self.content_box)
            root.add_widget(nav)

            self._init_ar_view()
            self._init_steps_view()

        except Exception as exc:
            db = _get_db()
            if db: db.log_error("NavigationScreen", "__init__", exc)
            traceback.print_exc()

    def _init_ar_view(self):
        # ── Full-screen cam card (AR-only mode) ──────────────────────────────
        self.cam_card = FloatLayout()
        self.cam_view = Image(allow_stretch=True, keep_ratio=False)
        self.cam_card.add_widget(self.cam_view)

        # ── Compact cam card (mixed AR + Steps mode) ─────────────────────────
        self.cam_card_small = RelativeLayout(
            size_hint=(1, None), height=dp(200),
        )
        with self.cam_card_small.canvas.before:
            self._cam_small_bg_color = Color(0.05, 0.05, 0.12, 0)
            self._cam_small_bg = Rectangle(
                pos=self.cam_card_small.pos,
                size=self.cam_card_small.size,
            )
        self.cam_card_small.bind(
            pos =lambda w, v: setattr(self._cam_small_bg, "pos",  v),
            size=lambda w, v: setattr(self._cam_small_bg, "size", v),
        )

        self.cam_view_small = Image(
            allow_stretch=True, keep_ratio=False,
            size_hint=(1, 1),
            color=(1, 1, 1, 0),
        )
        self.cam_card_small.add_widget(self.cam_view_small)

        # ── AR info panel ─────────────────────────────────────────────────────
        self._ar_info_panel = MDBoxLayout(
            orientation="vertical",
            size_hint=(1, 1),
            padding=[dp(6), dp(4)],
            spacing=dp(2),
            opacity=1,
        )
        self._ar_compass_lbl = MDLabel(
            text="↑", font_size=sp(34), bold=True, halign="center",
            theme_text_color="Custom", text_color=(0.3, 0.9, 0.5, 1),
            size_hint_y=None, height=dp(56),
        )
        self._ar_dist_big = MDLabel(
            text="— m", font_size=sp(15), bold=True, halign="center",
            theme_text_color="Custom", text_color=(1.0, 0.88, 0.3, 1),
            size_hint_y=None, height=dp(24),
        )
        self._ar_bear_lbl = MDLabel(
            text="Bearing: —", font_size=sp(6.5), halign="center",
            theme_text_color="Custom", text_color=(0.75, 0.65, 1.0, 1),
            size_hint_y=None, height=dp(13),
        )
        self._ar_status_lbl = MDLabel(
            text="📍 GPS locating…", font_size=sp(6), halign="center",
            theme_text_color="Custom", text_color=(0.55, 0.75, 0.55, 1),
            size_hint_y=None, height=dp(10),
        )
        self._ar_instr_lbl = MDLabel(text="")   # orphan — not in layout

        for w in [self._ar_compass_lbl, self._ar_dist_big,
                  self._ar_bear_lbl, self._ar_status_lbl]:
            self._ar_info_panel.add_widget(w)
        self.cam_card_small.add_widget(self._ar_info_panel)

        # ── GPS info overlay ──────────────────────────────────────────────────
        self.gps_overlay = MDBoxLayout(
            orientation="vertical",
            size_hint=(None, None),
            width=dp(148), height=dp(50),
            pos_hint={"right": 1, "top": 1},
            padding=[dp(4), dp(3)],
            spacing=dp(0),
            md_bg_color=(0.05, 0.05, 0.15, 0.82),
        )
        self._gps_src_lbl  = MDLabel(text="SRC: —", font_size=sp(5.5),
                                      theme_text_color="Custom",
                                      text_color=(0.55, 0.85, 1.0, 1),
                                      size_hint_y=None, height=dp(11))
        self._gps_dst_lbl  = MDLabel(text="DST: —", font_size=sp(5.5),
                                      theme_text_color="Custom",
                                      text_color=(0.55, 1.0, 0.65, 1),
                                      size_hint_y=None, height=dp(11))
        self._gps_dist_lbl = MDLabel(text="Dist: —", font_size=sp(5.5),
                                      theme_text_color="Custom",
                                      text_color=(1.0, 0.88, 0.3, 1),
                                      bold=True,
                                      size_hint_y=None, height=dp(11))
        self._gps_bear_lbl = MDLabel(text="Bearing: —", font_size=sp(5.5),
                                      theme_text_color="Custom",
                                      text_color=(0.85, 0.75, 1.0, 1),
                                      size_hint_y=None, height=dp(11))
        self._gps_user_lbl = MDLabel(text="You: GPS locating…")  # orphan

        for w in [self._gps_src_lbl, self._gps_dst_lbl,
                  self._gps_dist_lbl, self._gps_bear_lbl]:
            self.gps_overlay.add_widget(w)
        self.gps_overlay.opacity = 0
        self.cam_card.add_widget(self.gps_overlay)

    def _init_steps_view(self):
        self.steps_card = MDCard(
            orientation="vertical",
            md_bg_color=(0.97, 0.97, 1.0, 1), radius=0,
        )
        hdr = MDBoxLayout(
            size_hint=(1, None), height=dp(32),
            md_bg_color=(0.10, 0.30, 0.70, 1), padding=[dp(10), 0],
        )
        hdr.add_widget(MDLabel(
            text="Navigation Steps", font_size=sp(11), bold=True,
            theme_text_color="Custom", text_color=(1, 1, 1, 1),
            halign="left", valign="middle",
        ))
        self.steps_card.add_widget(hdr)
        self.scroll = ScrollView(
            bar_width=dp(5),
            bar_color=[0.15, 0.45, 0.90, 1.0],
            bar_inactive_color=[0.15, 0.45, 0.90, 0.35],
            scroll_type=["bars", "content"],
        )
        self.steps_list = MDBoxLayout(
            orientation="vertical", adaptive_height=True,
            padding=[dp(8), dp(8)], spacing=dp(6),
        )
        self.scroll.add_widget(self.steps_list)
        self.steps_card.add_widget(self.scroll)

    # ─────────────────────────────────────────────────────────────────────────

    def on_enter(self):
        # ── Request permissions when Activity window is VISIBLE ──────────────
        # This is the correct time on Android 6+ (Marshmallow and above).
        # Requesting earlier (at module load) may silently drop the dialog.
        if _IS_ANDROID:
            try:
                from android.permissions import (       # type: ignore
                    request_permissions, check_permission, Permission
                )
                global _CAMERA_GRANTED, _LOCATION_GRANTED

                def _perm_cb(perms, grants):
                    global _CAMERA_GRANTED, _LOCATION_GRANTED
                    cam_just_granted = False
                    for p, g in zip(perms, grants):
                        if p == Permission.CAMERA:
                            # Detect first-time grant so we can restart camera
                            cam_just_granted = (not _CAMERA_GRANTED) and g
                            _CAMERA_GRANTED = g
                            print(f"[Nav] on_enter CAMERA: {'GRANTED' if g else 'DENIED'}")
                        if p in (Permission.ACCESS_FINE_LOCATION,
                                 Permission.ACCESS_COARSE_LOCATION):
                            _LOCATION_GRANTED = _LOCATION_GRANTED or g
                    # FIX: if the user just granted the camera permission for the
                    # first time, restart the camera — the previous _start_camera()
                    # call skipped it because _CAMERA_GRANTED was still False.
                    if cam_just_granted:
                        print("[Nav] Camera newly granted — starting Kivy camera")
                        if hasattr(self, "_cam_target_view") and self._cam_target_view:
                            Clock.schedule_once(
                                lambda dt: self._start_kivy_camera(self._cam_target_view),
                                0.3
                            )
                        elif hasattr(self, "cam_view"):
                            Clock.schedule_once(
                                lambda dt: self._start_kivy_camera(self.cam_view),
                                0.3
                            )

                perms_needed = []
                if not check_permission(Permission.CAMERA):
                    perms_needed.append(Permission.CAMERA)
                if not check_permission(Permission.ACCESS_FINE_LOCATION):
                    perms_needed.append(Permission.ACCESS_FINE_LOCATION)
                if not check_permission(Permission.ACCESS_COARSE_LOCATION):
                    perms_needed.append(Permission.ACCESS_COARSE_LOCATION)

                if perms_needed:
                    request_permissions(perms_needed, _perm_cb)
                else:
                    _CAMERA_GRANTED   = True
                    _LOCATION_GRANTED = True
            except Exception as _pe:
                print(f"[Nav] on_enter permission request error: {_pe}")

        try:
            self.controller = NavigationController()
        except Exception as exc:
            print(f"[Nav] NavigationController init error: {exc}")
            self.controller = None
        try:
            self._room_data = _build_room_data()
        except Exception as exc:
            print(f"[Nav] _build_room_data error: {exc}")
            self._room_data = {}
        try:
            self._cam_small         = False
            self._first_frame       = True
            self._cam_target_view   = None
            self._active_ar_wrapper = None
        except Exception:
            pass
        try:
            _excel_path = _find_excel()
            if _excel_path:
                reload_campus_gps(_excel_path)
        except Exception as exc:
            print(f"[Nav] GPS reload error: {exc}")
        try:
            if hasattr(self, "src_btn"):
                self.src_btn.on_tap = lambda: self._open_popup(self.src_btn)
                self.src_btn.set_value("")
            if hasattr(self, "dst_btn"):
                self.dst_btn.on_tap = lambda: self._open_popup(self.dst_btn)
                self.dst_btn.set_value("")
        except Exception as exc:
            print(f"[Nav] on_enter btn setup error: {exc}")
        try:
            if hasattr(self, "hud"):
                self.hud.opacity = 0
            if hasattr(self, "content_box"):
                self.content_box.clear_widgets()
            if hasattr(self, "err_lbl"):
                self.err_lbl.text = ""
        except Exception as exc:
            print(f"[Nav] on_enter UI reset error: {exc}")

    def on_leave(self):
        try:
            from kivy.clock import Clock as _Clock
            _Clock.unschedule(self._update_frame)
            _Clock.unschedule(self._update_frame_mixed)
            _Clock.unschedule(self._update_kivy_ar)
            _Clock.unschedule(self._refresh_gps_overlay)
        except Exception as exc:
            print(f"[Nav] on_leave Clock.unschedule error: {exc}")
        try:
            self._stop_camera()
        except Exception as exc:
            print(f"[Nav] on_leave _stop_camera error: {exc}")
        try:
            self._restore_scroll()
        except Exception as exc:
            print(f"[Nav] on_leave _restore_scroll error: {exc}")
        try:
            if hasattr(self, "_active_popup") and self._active_popup:
                self._active_popup.dismiss()
        except Exception:
            pass
        try:
            self.controller = None
        except Exception:
            pass

    # ─────────────────────────────────────────────────────────────────────────

    def _open_popup(self, btn: NavDropBtn):
        if self._active_popup:
            try:
                self._active_popup.dismiss()
            except Exception:
                pass

        def on_pick(val: str):
            btn.set_value(val)
            self.err_lbl.text = ""

        popup = NavPopup(
            caller=btn,
            room_data=self._room_data,
            on_pick=on_pick,
        )
        self._active_popup = popup
        popup.open()

    # ─────────────────────────────────────────────────────────────────────────

    @staticmethod
    def _is_ar_location(val: str) -> bool:
        return val.strip() in AR_LOCATIONS_SET or val.strip().upper() in CAMPUS_BUILDINGS_SET

    def start_nav(self, *_):
        try:
            src = self.src_btn.text.strip() if hasattr(self, "src_btn") else ""
            dst = self.dst_btn.text.strip() if hasattr(self, "dst_btn") else ""
        except Exception as exc:
            print(f"[Nav] start_nav btn access error: {exc}")
            return

        if not src or not dst:
            if hasattr(self, "err_lbl"):
                self.err_lbl.text = "Please select both Source and Destination"
            return
        if src == dst:
            if hasattr(self, "err_lbl"):
                self.err_lbl.text = "Source and Destination must be different"
            return
        if hasattr(self, "err_lbl"):
            self.err_lbl.text = ""

        if not self.controller:
            if hasattr(self, "err_lbl"):
                self.err_lbl.text = "Navigation not ready. Please go back and retry."
            return

        try:
            ok, msg = self.controller.start_navigation(src, dst)
        except Exception as _e:
            traceback.print_exc()
            if hasattr(self, "err_lbl"):
                self.err_lbl.text = f"Navigation error: {_e}"
            return
        if not ok:
            if hasattr(self, "err_lbl"):
                self.err_lbl.text = f"Error: {msg}"
            return

        def _clean(v): return v.split("|", 1)[-1] if "|" in v else v
        src_disp = _clean(src)
        dst_disp = _clean(dst)

        try:
            if hasattr(self, "hud"):
                self.hud.opacity = 1
            self._stop_camera()
            if hasattr(self, "cam_card"):
                self._detach_widget(self.cam_card)
            if hasattr(self, "cam_card_small"):
                self._detach_widget(self.cam_card_small)
            self._restore_scroll()
            if hasattr(self, "content_box"):
                self.content_box.clear_widgets()
        except Exception as exc:
            print(f"[Nav] start_nav pre-branch UI error: {exc}")

        src_is_ar = self._is_ar_location(src)
        dst_is_ar = self._is_ar_location(dst)
        ctrl_wants_ar = self.controller.ar_overlay_active

        if src_is_ar and dst_is_ar:
            # ── Branch A: full-screen AR ──────────────────────────────────────
            try:
                self.instr_lbl.text      = ""
                self.gps_overlay.opacity = 0
                self._attach_gps_overlay(self.cam_card)
                self.content_box.add_widget(self.cam_card)
                self._populate_gps_overlay()
                self._start_camera(small=False)
            except Exception as _e:
                traceback.print_exc()
                self.err_lbl.text = f"AR start error: {_e}"

        elif ctrl_wants_ar:
            # ── Branch B: AR + Steps ──────────────────────────────────────────
            try:
                self.instr_lbl.text = ""

                ar_campus_key = ""
                if dst_is_ar:
                    _upper = dst.strip().upper()
                    if _upper in _CAMPUS_NAME_MAP:
                        ar_campus_key = _CAMPUS_NAME_MAP[_upper]
                    elif "|" in _upper:
                        ar_campus_key = _CAMPUS_NAME_MAP.get(
                            _upper.split("|", 1)[0].strip(), "")
                if not ar_campus_key and src_is_ar:
                    _upper = src.strip().upper()
                    if _upper in _CAMPUS_NAME_MAP:
                        ar_campus_key = _CAMPUS_NAME_MAP[_upper]
                    elif "|" in _upper:
                        ar_campus_key = _CAMPUS_NAME_MAP.get(
                            _upper.split("|", 1)[0].strip(), "")
                if not ar_campus_key:
                    ar_campus_key = (self.controller._dst_campus_key or
                                     self.controller._src_campus_key)

                if ar_campus_key:
                    self.controller.enable_ar_overlay(ar_campus_key)

                combined = BoxLayout(orientation="vertical", size_hint=(1, 1))

                ar_wrapper = RelativeLayout(
                    size_hint=(1, None),
                    height=max(dp(200), Window.height * 0.38),
                )
                with ar_wrapper.canvas.before:
                    Color(0.05, 0.05, 0.12, 1)
                    self._ar_wrapper_bg = Rectangle(
                        pos=ar_wrapper.pos, size=ar_wrapper.size)
                ar_wrapper.bind(
                    pos =lambda w, v: setattr(self._ar_wrapper_bg, "pos",  v),
                    size=lambda w, v: setattr(self._ar_wrapper_bg, "size", v),
                )
                self._detach_widget(self.cam_view_small)
                self.cam_view_small.size_hint = (1, 1)
                ar_wrapper.add_widget(self.cam_view_small)
                self._detach_widget(self.gps_overlay)
                ar_wrapper.add_widget(self.gps_overlay)
                self._detach_widget(self._ar_info_panel)
                ar_wrapper.add_widget(self._ar_info_panel)
                self._active_ar_wrapper = ar_wrapper
                combined.add_widget(ar_wrapper)

                step_header = MDBoxLayout(
                    size_hint=(1, None), height=dp(28),
                    md_bg_color=(0.10, 0.30, 0.70, 1),
                    padding=[dp(10), 0],
                )
                step_header.add_widget(MDLabel(
                    text="Navigation Steps", font_size=sp(9.5), bold=True,
                    theme_text_color="Custom", text_color=(1, 1, 1, 1),
                    halign="left", valign="middle",
                ))
                combined.add_widget(step_header)

                self._detach_widget(self.scroll)
                self.steps_list.clear_widgets()
                steps = self.controller.get_all_instructions_text()
                for idx, step in enumerate(steps, 1):
                    self.steps_list.add_widget(_make_step_card(idx, step))
                combined.add_widget(self.scroll)

                self.content_box.add_widget(combined)
                self._populate_gps_overlay()
                self.gps_overlay.opacity = 0
                self._start_camera_view(self.cam_view_small)
            except Exception as _e:
                traceback.print_exc()
                self.err_lbl.text = f"AR+Steps start error: {_e}"

        else:
            # ── Branch C: Steps only ──────────────────────────────────────────
            try:
                self._populate_steps(src_disp, dst_disp)
                self.content_box.add_widget(self.steps_card)
            except Exception as _e:
                traceback.print_exc()
                self.err_lbl.text = f"Steps error: {_e}"

    def _populate_steps(self, src_disp: str, dst_disp: str):
        self._restore_scroll()
        self.steps_list.clear_widgets()
        steps = self.controller.get_all_instructions_text()
        self.instr_lbl.text = f"{src_disp}  >>  {dst_disp}  ({len(steps)} steps)"
        for idx, step in enumerate(steps, 1):
            self.steps_list.add_widget(_make_step_card(idx, step))

    # ─────────────────────────────────────────────────────────────────────────

    def _populate_gps_overlay(self):
        if not self.controller:
            return
        s = self.controller.get_navigation_summary()
        src_gps  = s.get("src_gps",  "—")
        dst_gps  = s.get("dst_gps",  "—")
        dist_m   = s.get("static_distance_m")
        bearing  = s.get("live_bearing_deg")
        compass  = s.get("compass_direction", "")

        src_lbl = f"SRC: {src_gps}" if src_gps and src_gps != "—" else "SRC: indoor start"
        dst_lbl = f"DST: {dst_gps}" if dst_gps and dst_gps != "—" else "DST: —"

        dist_str = (f"{dist_m/1000:.2f} km" if dist_m and dist_m >= 1000
                    else f"{dist_m:.1f} m"   if dist_m is not None else "locating...")
        bear_str = (f"{bearing:.1f}°  {compass}" if bearing is not None else "locating...")

        self._gps_src_lbl.text  = src_lbl
        self._gps_dst_lbl.text  = dst_lbl
        self._gps_dist_lbl.text = f"Dist: {dist_str}"
        self._gps_bear_lbl.text = f"Bearing: {bear_str}"
        self._gps_user_lbl.text = "You: GPS locating…"
        self._update_kivy_ar()

    def _refresh_gps_overlay(self, dt=0):
        if not self.controller:
            return
        try:
            s = self.controller.get_navigation_summary()
        except Exception:
            return
        if not s.get("active"):
            return
        if not s.get("is_inter_building") and not s.get("ar_overlay_active"):
            return

        dist_live = s.get("live_distance_m")
        compass   = s.get("compass_direction", "")
        bearing   = s.get("live_bearing_deg")
        gps_fixed = s.get("gps_fixed", False)

        if dist_live is not None:
            dist_str = (f"{dist_live/1000:.2f} km" if dist_live >= 1000
                        else f"{dist_live:.1f} m")
            suffix   = "" if gps_fixed else " (est.)"
            self._gps_dist_lbl.text = f"Dist: {dist_str}{suffix}"

        if bearing is not None:
            self._gps_bear_lbl.text = f"Bearing: {bearing:.1f}°  {compass}"

        ctrl = self.controller
        if ctrl.gps_lat is not None:
            self._gps_user_lbl.text = (
                f"You: {ctrl.gps_lat:.6f}, {ctrl.gps_lon:.6f}"
            )
        if self._ar_info_panel.opacity > 0:
            self._update_kivy_ar()

    # ─────────────────────────────────────────────────────────────────────────
    #  ANDROID CAMERA FIX — Core camera methods
    # ─────────────────────────────────────────────────────────────────────────

    def _start_kivy_camera(self, target_image_widget, small: bool = False):
        """
        Start camera using Kivy's built-in Camera widget.
        This is the CORRECT approach for Android — uses the Camera2 API
        via Java bindings, respects Android 15 permission model.

        CRITICAL FIX (v3):
        The KivyCamera widget MUST be added to the widget tree (even off-screen)
        before Camera2 initialises. Without add_widget(), the Java Camera2
        backend never opens the hardware and the texture callback fires with
        None forever — resulting in a permanently black/blank camera view.

        Fix: add the widget to Window with size=(1,1) and opacity=0.
        It is invisible but present in the tree, which is all Camera2 needs.
        We bind its texture to our visible Image widget for AR compositing.
        """
        if not KIVY_CAMERA_AVAILABLE:
            print("[Nav] Kivy Camera widget not available — using AR panel only")
            return False

        if not _CAMERA_GRANTED:
            print("[Nav] Camera permission not granted — Kivy cam skipped")
            return False

        try:
            # Remove any existing Kivy camera cleanly
            if self._kivy_cam is not None:
                try:
                    self._kivy_cam.play = False
                    Window.remove_widget(self._kivy_cam)   # remove from tree
                except Exception:
                    pass
                self._kivy_cam = None

            # Create Kivy Camera widget
            # size=(1,1) + opacity=0 → invisible but PRESENT in widget tree.
            # Camera2 requires a rendered widget to open the hardware.
            self._kivy_cam = KivyCamera(
                play=False,
                resolution=(640, 480),
                index=0,
                size_hint=(None, None),
                size=(1, 1),        # 1×1 pixel — off-screen, invisible
                pos=(-1, -1),       # off-screen position
                opacity=0,          # invisible — we display via texture binding
                allow_stretch=True,
                keep_ratio=False,
            )

            # ── CRITICAL FIX: must be in the widget tree for Camera2 to open ──
            Window.add_widget(self._kivy_cam)
            print("[Nav] KivyCamera added to Window (Camera2 will now initialise)")

            # Bind Kivy camera texture to our visible display Image widget
            def _on_texture(camera, texture):
                if texture is not None and target_image_widget:
                    target_image_widget.texture = texture
                    if getattr(self, "_first_frame", True):
                        target_image_widget.color = (1, 1, 1, 1)
                        self._first_frame = False

            self._kivy_cam.bind(texture=_on_texture)

            # Warm-up delay: Android hardware context switching needs ~500ms
            def _enable_camera(dt):
                try:
                    if self._kivy_cam:
                        self._kivy_cam.play = True
                        print("[Nav] Kivy Camera started (Android mode)")
                except Exception as e:
                    print(f"[Nav] Kivy Camera play error: {e}")

            Clock.schedule_once(_enable_camera, 0.5)
            Clock.schedule_interval(self._refresh_gps_overlay, 1.0)
            return True

        except Exception as e:
            print(f"[Nav] _start_kivy_camera error: {e}")
            traceback.print_exc()
            return False

    def _start_cv2_camera(self, small: bool = False):
        """
        Start camera using OpenCV VideoCapture.
        DESKTOP ONLY — unreliable on Android.

        Android camera failure root causes:
        1. cv2.VideoCapture on Android requires V4L2 (Video4Linux2) drivers.
           Android uses Camera2 API (Java layer), NOT V4L2.
        2. Buildozer packages opencv-python which is compiled for Linux x86.
           The Android ARM build lacks proper camera backend support.
        3. Even if a frame is returned, it is usually all-black because
           the Camera2 → V4L2 bridge is not set up in the APK environment.

        FIX: On Android, use _start_kivy_camera() instead.
              On desktop, this method works correctly.
        """
        if not CV2_AVAILABLE:
            return False

        if _IS_ANDROID:
            print("[Nav] ANDROID: cv2.VideoCapture skipped — use Kivy Camera instead")
            return False

        # Desktop: try multiple camera indices with delay
        import time

        target_view = (self.cam_view_small if small else self.cam_view)

        for cam_index in [0, 1, -1]:
            try:
                time.sleep(0.2)   # warm-up delay between attempts
                _cap = cv2.VideoCapture(cam_index)

                if not _cap.isOpened():
                    _cap.release()
                    continue

                # Set resolution explicitly to avoid buffer size mismatches
                _cap.set(cv2.CAP_PROP_FRAME_WIDTH,  640)
                _cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
                _cap.set(cv2.CAP_PROP_FPS, 30)

                # Warm-up: read and discard first few frames
                for _ in range(3):
                    _cap.read()

                # Verify we get real frames (not black)
                ret, test_frame = _cap.read()
                if not ret or test_frame is None:
                    _cap.release()
                    continue

                self.cap = _cap
                self._cam_small = small
                print(f"[Nav] cv2 camera opened at index {cam_index} (desktop mode)")

                if small:
                    Clock.schedule_interval(self._update_frame_mixed, 1 / 30)
                else:
                    Clock.schedule_interval(self._update_frame, 1 / 30)
                Clock.schedule_interval(self._refresh_gps_overlay, 1.0)
                return True

            except Exception as _e:
                print(f"[Nav] cv2.VideoCapture({cam_index}) error: {_e}")

        print("[Nav] cv2: all camera indices failed")
        return False

    def _start_camera(self, small: bool = False):
        """
        Main camera entry point.

        ANDROID CAMERA FIX:
        - Android 15 → use Kivy Camera widget (Camera2 API via Java)
        - Desktop/Linux → use cv2.VideoCapture (V4L2)
        - Fallback: Kivy AR compass panel (always works, no camera needed)

        The AR panel is shown IMMEDIATELY as primary UI.
        Camera feed overlays on top as a bonus visual layer.
        """
        self._stop_camera()
        self._cam_small   = small
        self._first_frame = True

        target_view = self.cam_view_small if small else self.cam_view

        # ── Step 1: Show AR panel immediately (primary, no camera needed) ────
        if not CV2_AVAILABLE or _IS_ANDROID:
            self._ar_info_panel.opacity = 1
            self._update_kivy_ar()
            Clock.schedule_interval(self._update_kivy_ar, 0.33)

        # ── Step 2: Activate background color for AR area ────────────────────
        if small and hasattr(self, "_cam_small_bg_color"):
            self._cam_small_bg_color.a = 1

        # ── Step 3: Try to start camera ───────────────────────────────────────
        cam_started = False

        if _IS_ANDROID:
            # ANDROID: use Kivy Camera widget
            if _CAMERA_GRANTED:
                cam_started = self._start_kivy_camera(target_view, small=small)
            else:
                print("[Nav] Camera permission not granted — AR compass panel only")
                self.err_lbl.text = "Camera permission required. Tap Settings to grant."
        else:
            # DESKTOP: use cv2
            cam_started = self._start_cv2_camera(small=small)

        if not cam_started:
            # Fallback: AR compass panel is already showing
            print("[Nav] Camera unavailable — showing Kivy AR compass panel")
            if not _IS_ANDROID:
                # On desktop, also start GPS refresh
                Clock.schedule_interval(self._refresh_gps_overlay, 1.0)
            if not small:
                # Full-screen: move AR panel to cam_card
                self._detach_widget(self._ar_info_panel)
                self.cam_card.add_widget(self._ar_info_panel)
            self._ar_info_panel.opacity = 1
            self._update_kivy_ar()
            Clock.schedule_interval(self._update_kivy_ar, 0.33)
            Clock.schedule_interval(self._refresh_gps_overlay, 1.0)

    def _start_camera_view(self, target_view):
        """
        Start mixed-mode camera (Branch B: AR + Steps).
        Shows AR panel immediately; camera overlays when available.
        """
        Clock.unschedule(self._update_frame)
        Clock.unschedule(self._update_frame_mixed)
        Clock.unschedule(self._update_kivy_ar)
        if self.cap:
            try: self.cap.release()
            except Exception: pass
            self.cap = None

        if hasattr(self, "_cam_small_bg_color"):
            self._cam_small_bg_color.a = 1

        # Show AR panel immediately
        self._ar_info_panel.opacity = 1
        self._update_kivy_ar()
        Clock.schedule_interval(self._update_kivy_ar, 0.33)

        self._cam_target_view = target_view
        self._cam_target_view.color = (1, 1, 1, 0)
        self._first_frame = True
        Clock.schedule_interval(self._refresh_gps_overlay, 1.0)

        # Try camera
        if _IS_ANDROID:
            if _CAMERA_GRANTED:
                self._start_kivy_camera(target_view, small=True)
        else:
            if CV2_AVAILABLE:
                self._start_cv2_camera(small=True)

    def _stop_camera(self):
        """Stop all camera providers and clean up textures."""
        try:
            Clock.unschedule(self._update_frame)
            Clock.unschedule(self._update_frame_mixed)
            Clock.unschedule(self._update_kivy_ar)
        except Exception:
            pass

        # Stop Kivy Camera — MUST remove from Window to fully release Camera2
        if self._kivy_cam is not None:
            try:
                self._kivy_cam.play = False
                # FIX: remove from Window tree so Camera2 releases the hardware.
                # Failing to do this causes "camera already in use" errors on
                # the next _start_kivy_camera() call (e.g. after going back and
                # re-entering the Navigation screen).
                Window.remove_widget(self._kivy_cam)
                print("[Nav] KivyCamera removed from Window (Camera2 released)")
            except Exception:
                pass
            self._kivy_cam = None

        # Stop cv2 capture
        if getattr(self, "cap", None):
            try:
                self.cap.release()
            except Exception:
                pass
            self.cap = None

        # Clear textures
        if hasattr(self, "cam_view"):
            try:
                self.cam_view.texture = None
            except Exception:
                pass
        if hasattr(self, "cam_view_small"):
            try:
                self.cam_view_small.texture = None
                self.cam_view_small.color   = (1, 1, 1, 0)
            except Exception:
                pass
        if hasattr(self, "_cam_small_bg_color"):
            try:
                self._cam_small_bg_color.a = 0
            except Exception:
                pass
        if hasattr(self, "_cam_target_view") and self._cam_target_view:
            try:
                self._cam_target_view.texture = None
                self._cam_target_view.color   = (1, 1, 1, 0)
                self._cam_target_view         = None
            except Exception:
                pass
        if hasattr(self, "gps_overlay"):
            try:
                self.gps_overlay.opacity = 0
            except Exception:
                pass
        if hasattr(self, "_ar_info_panel"):
            try:
                self._ar_info_panel.opacity = 0
            except Exception:
                pass
        try:
            if (hasattr(self, "_ar_info_panel") and hasattr(self, "cam_card_small") and
                    self._ar_info_panel.parent is not None and
                    self._ar_info_panel.parent is not self.cam_card_small):
                self._detach_widget(self._ar_info_panel)
                self.cam_card_small.add_widget(self._ar_info_panel)
        except Exception:
            pass
        self._first_frame = True

    def _update_frame(self, dt):
        """cv2 frame update — desktop only."""
        if not CV2_AVAILABLE or not (self.cap and self.cap.isOpened()):
            return
        ret, frame = self.cap.read()
        if not ret or frame is None:
            return
        self._refresh_gps_overlay()
        if self.controller:
            frame = self.controller.render_ar_frame(frame)
        buf = cv2.flip(frame, 0).tobytes()
        tex = Texture.create(
            size=(frame.shape[1], frame.shape[0]), colorfmt="bgr"
        )
        tex.blit_buffer(buf, colorfmt="bgr", bufferfmt="ubyte")
        if getattr(self, "_cam_small", False):
            if getattr(self, "_first_frame", True):
                self.cam_view_small.color = (1, 1, 1, 1)
                self._first_frame = False
            self.cam_view_small.texture = tex
        else:
            self.cam_view.texture = tex

    def _update_frame_mixed(self, dt):
        """cv2 mixed-mode frame update — desktop only."""
        if not CV2_AVAILABLE or not (self.cap and self.cap.isOpened()):
            return
        ret, frame = self.cap.read()
        if not ret or frame is None:
            return

        try:
            import numpy as np
            is_real_frame = bool(np.mean(frame) > 8)
        except ImportError:
            is_real_frame = True

        if is_real_frame:
            self._refresh_gps_overlay()
            if self.controller:
                frame = self.controller.render_ar_frame(frame)
            buf = cv2.flip(frame, 0).tobytes()
            tex = Texture.create(size=(frame.shape[1], frame.shape[0]), colorfmt="bgr")
            tex.blit_buffer(buf, colorfmt="bgr", bufferfmt="ubyte")
            target = getattr(self, "_cam_target_view", None)
            if target:
                if getattr(self, "_first_frame", True):
                    target.color   = (1, 1, 1, 1)
                    self._first_frame = False
                target.texture = tex
        else:
            if getattr(self, "_cam_target_view", None):
                self._cam_target_view.color = (1, 1, 1, 0)
            self._first_frame = True

    def _update_kivy_ar(self, dt=0):
        """
        Refresh the Kivy-drawn AR compass panel.
        This is the PRIMARY AR display — works without any camera.
        """
        if not self.controller:
            return
        try:
            s = self.controller.get_navigation_summary()
        except Exception:
            return
        bearing = s.get("live_bearing_deg")
        compass = s.get("compass_direction", "")
        gps_ok  = s.get("gps_fixed", False)

        dist_m = s.get("live_distance_m")
        if dist_m is None:
            dist_m = s.get("static_distance_m")

        if dist_m is not None:
            d_str  = f"{dist_m/1000:.2f} km" if dist_m >= 1000 else f"{int(dist_m)} m"
            suffix = "" if gps_ok else " (est.)"
            self._ar_dist_big.text       = d_str + suffix
            self._ar_dist_big.text_color = (
                (0.3, 0.95, 0.5, 1) if gps_ok else (1.0, 0.88, 0.3, 1)
            )
        else:
            self._ar_dist_big.text       = "Calculating…"
            self._ar_dist_big.text_color = (0.6, 0.8, 1.0, 1)

        _arrow_map = {
            "N": "↑", "NNE": "↑",  "NE": "↗", "ENE": "→", "E": "→",
            "ESE": "→", "SE": "↘", "SSE": "↓", "S": "↓",  "SSW": "↓",
            "SW": "↙", "WSW": "←", "W": "←",  "WNW": "←", "NW": "↖", "NNW": "↑",
        }
        arrow = _arrow_map.get(compass, "↑")
        self._ar_compass_lbl.text       = arrow
        self._ar_compass_lbl.text_color = (
            (0.3, 0.9, 0.5, 1) if gps_ok else (1.0, 0.78, 0.2, 1)
        )

        self._ar_bear_lbl.text = (
            f"Bearing  {bearing:.1f}°  {compass}"
            if bearing is not None else "Bearing: locating…"
        )

        ctrl = self.controller
        if ctrl.gps_lat is not None:
            self._ar_status_lbl.text = f"📍 {ctrl.gps_lat:.5f}, {ctrl.gps_lon:.5f}"
        else:
            self._ar_status_lbl.text = "📍 GPS locating…"

    # ─────────────────────────────────────────────────────────────────────────

    def _detach_widget(self, widget):
        try:
            if widget and widget.parent:
                widget.parent.remove_widget(widget)
        except Exception:
            pass

    def _attach_gps_overlay(self, target_card):
        self._detach_widget(self.gps_overlay)
        target_card.add_widget(self.gps_overlay)

    def _restore_scroll(self):
        try:
            if hasattr(self, "scroll") and hasattr(self, "steps_card"):
                if self.scroll.parent is self.steps_card:
                    pass
                else:
                    self._detach_widget(self.scroll)
                    self.steps_card.add_widget(self.scroll)
        except Exception:
            pass
        try:
            if (hasattr(self, "cam_view_small") and hasattr(self, "cam_card_small") and
                    self.cam_view_small.parent is not None and
                    self.cam_view_small.parent is not self.cam_card_small):
                self._detach_widget(self.cam_view_small)
                self.cam_card_small.add_widget(self.cam_view_small)
        except Exception:
            pass
        try:
            if (hasattr(self, "gps_overlay") and hasattr(self, "cam_card") and
                    self.gps_overlay.parent is not None and
                    self.gps_overlay.parent is not self.cam_card):
                self._detach_widget(self.gps_overlay)
                self.cam_card.add_widget(self.gps_overlay)
        except Exception:
            pass
        try:
            if (hasattr(self, "_ar_info_panel") and hasattr(self, "cam_card_small") and
                    self._ar_info_panel.parent is not None and
                    self._ar_info_panel.parent is not self.cam_card_small):
                self._detach_widget(self._ar_info_panel)
                self.cam_card_small.add_widget(self._ar_info_panel)
        except Exception:
            pass

    def _on_back(self):
        try:
            if hasattr(self, "hud") and self.hud.opacity == 1:
                self._stop_camera()
                self.hud.opacity = 0
                if hasattr(self, "content_box"):
                    self.content_box.clear_widgets()
                self._restore_scroll()
                if hasattr(self, "cam_card"):
                    self._attach_gps_overlay(self.cam_card)
                if hasattr(self, "err_lbl"):
                    self.err_lbl.text = ""
            else:
                self._safe_go_back()
        except Exception as exc:
            print(f"[Nav] _on_back error: {exc}")
            try:
                self._safe_go_back()
            except Exception:
                pass

    def _safe_go_back(self):
        try:
            if self.manager:
                self.manager.transition.direction = "right"
                self.manager.current = "dashboard"
        except Exception as exc:
            db = _get_db()
            if db:
                db.log_error("NavigationScreen", "_safe_go_back", exc)
            traceback.print_exc()
